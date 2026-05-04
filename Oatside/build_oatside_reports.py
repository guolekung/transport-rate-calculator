"""
Build Oatside -> P&G trip summary from raw GPS workbooks (sheet อุปกรณ์),
write Oatside_PG_Trip_Summary_By_Site.xlsx + HTML under TransportRateCalculator/reports/oatside-apr2026.

Run from repo root:
  python Oatside/build_oatside_reports.py

Billing rules are read from Oatside/oatside_config.json (auto-created with defaults if missing).
Override JSON: Oatside/oatside_billing_overrides.json (or env OATSIDE_OVERRIDES_JSON).

Optional env:
  OATSIDE_ORIGIN        = path to Oatside point xlsx (else newest Y.K.*Oatside*.xlsx in Oatside/)
  OATSIDE_DEST          = path to P&G point xlsx (else newest Y.K.*P&G* or *เวลล์โกล*.xlsx)
  OATSIDE_MAX_TRAVEL_H  = overrides config max_travel_h (max hours Origin_Out->Dest_In for pairing)
  OATSIDE_OVERRIDES_JSON = optional path to billing overrides JSON
"""

from __future__ import annotations

import html as html_module
import math
import os
import re
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path
import json
from typing import Any, Callable, Iterable

import openpyxl

BIGC_EXACT = frozenset({"71-5041", "71-5042"})
PLATE_HEAD = re.compile(r"^(\d{2}-\d{4})\b")
DETAIL_KEY = re.compile(r"^\d+\.\d+$")

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

@dataclass
class OatsideConfig:
    trip_rates: list[dict]
    one_trip_surcharge_pct: float
    min_trips_per_truck: int
    max_travel_h: float
    max_origin_chain_gap_h: float
    enable_origin_chain_merge: bool
    charge_min_trip_shortfall: bool
    use_origin_24h_fifty: bool
    use_origin_day_fifty: bool
    customer_idle_windows: list[CustomerIdleWindow]
    customer_no_work_ranges: list[tuple[date, date, str]]
    outbound_half_dest_dates: frozenset[date]
    long_dest_wait_midnight_fifty: bool
    long_dest_wait_midnight_min_h: float
    long_dest_wait_midnight_full_trip: bool
    highlight_origin_wait_h: float
    highlight_dest_wait_h: float
    manual_extra_trips: tuple[ManualExtraTrip, ...]
    manual_return_trips: tuple[ManualExtraTrip, ...]

@dataclass
class CustomerIdleWindow:
    """Hours at customer site excluded from customer dwell / 24h gap (e.g. factory parking)."""

    plate: str
    start: datetime
    end: datetime
    note: str = ""

    def overlaps_dest_interval(self, d_in: datetime, d_out: datetime) -> bool:
        return d_in < self.end and d_out > self.start

    def overlap_hours(self, d_in: datetime, d_out: datetime) -> float:
        a = max(d_in, self.start)
        b = min(d_out, self.end)
        if b <= a:
            return 0.0
        return (b - a).total_seconds() / 3600.0


@dataclass(frozen=True)
class ManualExtraTrip:
    """ลูกค้าตกลงเก็บเพิ่มแต่ไม่มีในไฟล์ GPS (เช่น P&G→Oatside)."""

    dest_date: date
    plate: str
    amount_baht: int
    note: str = ""



_DEFAULT_NO_WORK_RANGES: list[tuple[date, date, str]] = [
    (date(2026, 4, 23), date(2026, 4, 24), "customer no-work"),
    (date(2026, 4, 27), date(2026, 4, 28), "customer no-work"),
    (date(2026, 5, 1), date(2026, 5, 1), "customer no-work"),
]


def _recovery_dest_dates_from_no_work(ranges: list[tuple[date, date, str]]) -> frozenset[date]:
    """First calendar day after each no-work block ends (Dest_In date for first trip surcharge)."""
    return frozenset(b + timedelta(days=1) for _a, b, _n in ranges)


_DEFAULT_OUTBOUND_HALF_DATES: frozenset[date] = _recovery_dest_dates_from_no_work(_DEFAULT_NO_WORK_RANGES)



_DEFAULT_CONFIG = OatsideConfig(
    trip_rates=[
        {"from": "2026-04-12", "to": "2026-04-15", "rate_baht": 8000},
        {"rate_baht": 7500},
    ],
    one_trip_surcharge_pct=50.0,
    min_trips_per_truck=2,
    max_travel_h=48.0,
    max_origin_chain_gap_h=3.0,
    enable_origin_chain_merge=False,
    charge_min_trip_shortfall=False,
    use_origin_24h_fifty=True,
    use_origin_day_fifty=True,
    customer_idle_windows=[
        CustomerIdleWindow(
            plate="71-8967",
            start=datetime(2026, 4, 20, 14, 0, 0),
            end=datetime(2026, 4, 29, 17, 0, 0),
            note="Factory parked CONTEXT_LOG 90-91",
        ),
    ],
    customer_no_work_ranges=list(_DEFAULT_NO_WORK_RANGES),
    outbound_half_dest_dates=_DEFAULT_OUTBOUND_HALF_DATES,
    long_dest_wait_midnight_fifty=True,
    long_dest_wait_midnight_min_h=12.0,
    long_dest_wait_midnight_full_trip=True,
    highlight_origin_wait_h=8.0,
    highlight_dest_wait_h=8.0,
    manual_extra_trips=(),
    manual_return_trips=(),
)


_DEFAULT_CONFIG_JSON = {
    "version": 1,
    "_help": "แก้ไฟล์นี้เพื่อเปลี่ยนกฎการคิดเงิน แล้วรัน build_oatside_reports.py ใหม่",
    "trip_rates": [
        {"_note": "ช่วงวันที่พิเศษ (สงกรานต์ 2026)", "from": "2026-04-12", "to": "2026-04-15", "rate_baht": 8000},
        {"_note": "เรทปกติ (ใช้เมื่อไม่ตรงช่วงไหนข้างบน)", "rate_baht": 7500},
    ],
    "one_trip_surcharge_pct": 50,
    "min_trips_per_truck_per_day": 2,
    "max_travel_h": 48,
    "max_origin_chain_gap_h": 3,
    "_note_max_origin_chain_gap_h": "hours: max gap Origin_Out(prev)->Origin_In(next) for chain-merge; larger gap = new hub visit (only if enable_origin_chain_merge is true)",
    "enable_origin_chain_merge": False,
    "_note_enable_origin_chain_merge": "false = never merge multiple Origin rows before one Dest (disables merge_chained_origin_pairs entirely)",
    "use_origin_day_fifty": True,
    "_note_use_origin_day_fifty": "true (default) = 50pct เมื่อ 1 เที่ยวต่อวันงาน (Origin_In calendar day) | ข้ามคืนไม่แตกวัน | วันไม่มี Origin ไม่นับ | มีผลก่อน use_origin_24h_fifty",
    "use_origin_24h_fifty": True,
    "_note_use_origin_24h_fifty": "true = 50pct downtime from rolling 24h windows anchored at each trip Origin_In chain; false = legacy Dest_In calendar day (1 trip => +50pct)",
    "customer_idle_windows": [
        {
            "_note": "71-8967 P&G factory parking — customer-irrelevant dwell (CONTEXT_LOG Session #90–91)",
            "plate": "71-8967",
            "start": "2026-04-20 14:00:00",
            "end": "2026-04-29 17:00:00",
            "note": "Parked at customer — clip dest wait from Daily_Time / gap vs 24h",
        },
    ],
    "customer_no_work": [
        {"from": "2026-04-23", "to": "2026-04-24", "note": "customer no-work"},
        {"from": "2026-04-27", "to": "2026-04-28", "note": "customer no-work"},
        {"from": "2026-05-01", "to": "2026-05-01", "note": "customer no-work"}
    ],
    "long_dest_wait_midnight_fifty": True,
    "long_dest_wait_midnight_min_h": 12,
    "long_dest_wait_midnight_full_trip": True,
    "_note_long_dest_wait_midnight_full": "true = charge full 1-trip rate on dest_date when midnight dwell rule fires; false = charge one_trip_surcharge_pct of rate",
    "highlight_origin_wait_h": 8,
    "highlight_dest_wait_h": 8,
    "manual_extra_trips": [],
    "_note_manual_extra_trips": "เที่ยวเพิ่มที่ไม่มีใน GPS — ตัวอย่าง: {\"dest_date\": \"2026-04-22\", \"plate\": \"72-1217\", \"amount_baht\": 7500, \"note\": \"P&G->Oatside\"}",
    "_note_long_dest_wait_midnight": "If Dest_In and Dest_Out cross midnight and dwell >= min_h, add surcharge by dest_date when no fifty row yet (origin_day mode gap)",
    "_note_outbound_half": "If outbound_half_dest_dates omitted, recovery = day after each no-work block end; surcharge 50pct on first matched trip that Dest_In day",
    "charge_min_trip_shortfall": False,
    "_note_charge_min_trip_shortfall": "ถ้า false = ไม่เก็บเงินค่าชดเชยเที่ยวขาด (min trips) ในรายงานลูกค้า — ใช้ชาร์จ % วันละ 1 เที่ยวแทน | true = เก็บทั้งค่าชดเชย + % ตามเดิม",
}


def _config_path() -> Path:
    return _oatside_dir() / "oatside_config.json"


def load_oatside_config() -> OatsideConfig:
    """Load billing config from oatside_config.json; fall back to built-in defaults."""
    path = _config_path()
    if not path.is_file():
        # Write a default config so โอ can see/edit it
        try:
            path.write_text(json.dumps(_DEFAULT_CONFIG_JSON, ensure_ascii=False, indent=2), encoding="utf-8")
            print(f"[INFO] สร้าง config เริ่มต้น: {path}")
        except OSError:
            pass
        return _DEFAULT_CONFIG

    try:
        raw = json.loads(path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        print(f"[WARN] อ่าน {path} ไม่ได้ — ใช้ค่า default แทน")
        return _DEFAULT_CONFIG

    if not isinstance(raw, dict):
        return _DEFAULT_CONFIG

    trip_rates = raw.get("trip_rates", _DEFAULT_CONFIG.trip_rates)
    if not isinstance(trip_rates, list) or not trip_rates:
        trip_rates = _DEFAULT_CONFIG.trip_rates

    surcharge_pct = float(raw.get("one_trip_surcharge_pct", _DEFAULT_CONFIG.one_trip_surcharge_pct))
    min_trips = int(raw.get("min_trips_per_truck_per_day", _DEFAULT_CONFIG.min_trips_per_truck))

    env_travel = os.environ.get("OATSIDE_MAX_TRAVEL_H")
    if env_travel:
        try:
            max_travel = float(env_travel)
        except ValueError:
            max_travel = float(raw.get("max_travel_h", _DEFAULT_CONFIG.max_travel_h))
    else:
        max_travel = float(raw.get("max_travel_h", _DEFAULT_CONFIG.max_travel_h))

    charge_sf = bool(raw.get("charge_min_trip_shortfall", False))
    try:
        gap_h = float(raw.get("max_origin_chain_gap_h", _DEFAULT_CONFIG.max_origin_chain_gap_h))
    except (TypeError, ValueError):
        gap_h = float(_DEFAULT_CONFIG.max_origin_chain_gap_h)

    chain_merge = bool(raw.get("enable_origin_chain_merge", _DEFAULT_CONFIG.enable_origin_chain_merge))
    use_o24 = bool(raw.get("use_origin_24h_fifty", _DEFAULT_CONFIG.use_origin_24h_fifty))
    use_o_day = bool(raw.get("use_origin_day_fifty", _DEFAULT_CONFIG.use_origin_day_fifty))
    if "customer_idle_windows" not in raw:
        idle_raw = _DEFAULT_CONFIG_JSON["customer_idle_windows"]
    else:
        idle_raw = raw.get("customer_idle_windows") or []
    idle_wins: list[CustomerIdleWindow] = []
    if isinstance(idle_raw, list):
        for w in idle_raw:
            if not isinstance(w, dict):
                continue
            pl = str(w.get("plate", "")).strip()
            st = _parse_dt(w.get("start"))
            en = _parse_dt(w.get("end"))
            if not pl or not st or not en or en <= st:
                continue
            note = str(w.get("note", "")).strip()
            idle_wins.append(CustomerIdleWindow(plate=pl, start=st, end=en, note=note))

    if "customer_no_work" not in raw:
        nwr = list(_DEFAULT_NO_WORK_RANGES)
    else:
        nwr = _parse_no_work_entries(raw.get("customer_no_work"))
        if not nwr:
            nwr = list(_DEFAULT_NO_WORK_RANGES)
    ohd = _parse_date_set(raw.get("outbound_half_dest_dates"))
    if not ohd:
        ohd = frozenset(_recovery_dest_dates_from_no_work(nwr))

    manual_list: list[ManualExtraTrip] = []
    raw_mx = raw.get("manual_extra_trips")
    if isinstance(raw_mx, list):
        for e in raw_mx:
            if not isinstance(e, dict):
                continue
            ds = str(e.get("dest_date", "")).strip()[:10]
            pl = str(e.get("plate", "")).strip()
            try:
                amt = int(e.get("amount_baht", 0) or 0)
            except (TypeError, ValueError):
                amt = 0
            note = str(e.get("note", "")).strip()
            if len(ds) < 10 or not pl or amt <= 0:
                continue
            try:
                dd = datetime.strptime(ds, "%Y-%m-%d").date()
            except ValueError:
                continue
            manual_list.append(ManualExtraTrip(dest_date=dd, plate=pl, amount_baht=amt, note=note))

    return_list: list[ManualExtraTrip] = []
    raw_rt = raw.get("manual_return_trips")
    if isinstance(raw_rt, list):
        for e in raw_rt:
            if not isinstance(e, dict):
                continue
            ds = str(e.get("dest_date", "")).strip()[:10]
            pl = str(e.get("plate", "")).strip()
            try:
                amt = int(e.get("amount_baht", 0) or 0)
            except (TypeError, ValueError):
                amt = 0
            note = str(e.get("note", "")).strip()
            if len(ds) < 10 or not pl or amt <= 0:
                continue
            try:
                dd = datetime.strptime(ds, "%Y-%m-%d").date()
            except ValueError:
                continue
            return_list.append(ManualExtraTrip(dest_date=dd, plate=pl, amount_baht=amt, note=note))

    return OatsideConfig(
        trip_rates=trip_rates,
        one_trip_surcharge_pct=surcharge_pct,
        min_trips_per_truck=min_trips,
        max_travel_h=max_travel,
        max_origin_chain_gap_h=gap_h,
        enable_origin_chain_merge=chain_merge,
        charge_min_trip_shortfall=charge_sf,
        use_origin_24h_fifty=use_o24,
        use_origin_day_fifty=use_o_day,
        customer_idle_windows=idle_wins,
        customer_no_work_ranges=nwr,
        outbound_half_dest_dates=ohd,
        long_dest_wait_midnight_fifty=bool(
            raw.get("long_dest_wait_midnight_fifty", _DEFAULT_CONFIG.long_dest_wait_midnight_fifty)
        ),
        long_dest_wait_midnight_min_h=float(
            raw.get("long_dest_wait_midnight_min_h", _DEFAULT_CONFIG.long_dest_wait_midnight_min_h)
        ),
        long_dest_wait_midnight_full_trip=bool(
            raw.get("long_dest_wait_midnight_full_trip", _DEFAULT_CONFIG.long_dest_wait_midnight_full_trip)
        ),
        highlight_origin_wait_h=float(
            raw.get("highlight_origin_wait_h", _DEFAULT_CONFIG.highlight_origin_wait_h)
        ),
        highlight_dest_wait_h=float(
            raw.get("highlight_dest_wait_h", _DEFAULT_CONFIG.highlight_dest_wait_h)
        ),
        manual_extra_trips=tuple(manual_list),
        manual_return_trips=tuple(return_list),
    )


def trip_rate_baht(d: date, cfg: OatsideConfig) -> int:
    """Look up trip rate for a given Dest_In date using config rules (first match wins)."""
    for rule in cfg.trip_rates:
        frm = rule.get("from")
        to = rule.get("to")
        rate = rule.get("rate_baht")
        if rate is None:
            continue
        if frm and to:
            try:
                d_from = datetime.strptime(str(frm), "%Y-%m-%d").date()
                d_to = datetime.strptime(str(to), "%Y-%m-%d").date()
                if d_from <= d <= d_to:
                    return int(rate)
            except ValueError:
                continue
        else:
            return int(rate)
    return 7500


def config_rate_summary(cfg: OatsideConfig) -> str:
    """Human-readable summary of rate rules for subtitles/logs."""
    parts = []
    for rule in cfg.trip_rates:
        rate = rule.get("rate_baht")
        frm = rule.get("from")
        to = rule.get("to")
        if rate is None:
            continue
        if frm and to:
            parts.append(f"{frm}–{to}={rate:,}")
        else:
            parts.append(f"ปกติ={rate:,}")
    return " / ".join(parts) if parts else "7,500"


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------

@dataclass
class Leg:
    row_no: str
    plate: str
    device: str
    t_in: datetime
    t_out: datetime


@dataclass
class Trip:
    plate: str
    site: str
    device: str
    o_row: str
    d_row: str
    o_in: datetime
    o_out: datetime
    d_in: datetime
    d_out: datetime
    origin_wait_h: float
    travel_h: float
    dest_wait_h: float
    total_cycle_h: float
    origin_date: date
    dest_date: date
    trip_date: date
    travel_flag: str | None


# ---------------------------------------------------------------------------
# Path helpers
# ---------------------------------------------------------------------------

def _root() -> Path:
    here = Path(__file__).resolve().parent
    return here.parent if (here.parent / "TransportRateCalculator").is_dir() else here


def _oatside_dir() -> Path:
    return Path(__file__).resolve().parent


def _overrides_json_path() -> Path:
    v = os.environ.get("OATSIDE_OVERRIDES_JSON")
    if v:
        return Path(v)
    return _oatside_dir() / "oatside_billing_overrides.json"


# ---------------------------------------------------------------------------
# Overrides loader
# ---------------------------------------------------------------------------

def load_billing_overrides() -> dict[tuple[str, date], dict[str, Any]]:
    """Load manual billing actions keyed by (plate, dest_date).

    JSON shape (UTF-8):
      {"version": 1, "entries": [
        {"dest_date": "2026-04-14", "plate": "71-6802", "action": "exclude_50", "note": "..."},
        {"dest_date": "2026-04-20", "plate": "71-6001", "action": "include_50", "note": "..."}
      ]}

    - exclude_50: do not charge 50% even if auto rule would (exactly 1 matched trip that Dest_In day).
    - include_50: charge 50% of one trip rate that day even if auto rule would not (e.g. 2+ trips).
    """
    path = _overrides_json_path()
    out: dict[tuple[str, date], dict[str, Any]] = {}
    if not path.is_file():
        return out
    try:
        raw = json.loads(path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return out
    entries = raw.get("entries") if isinstance(raw, dict) else None
    if not isinstance(entries, list):
        return out
    for e in entries:
        if not isinstance(e, dict):
            continue
        ds = str(e.get("dest_date", "")).strip()[:10]
        plate = str(e.get("plate", "")).strip()
        action = str(e.get("action", "")).strip()
        note = str(e.get("note", "")).strip()
        if not ds or not plate or action not in ("exclude_50", "include_50"):
            continue
        try:
            d = datetime.strptime(ds, "%Y-%m-%d").date()
        except ValueError:
            continue
        out[(plate, d)] = {"action": action, "note": note}
    return out


# ---------------------------------------------------------------------------
# File discovery
# ---------------------------------------------------------------------------

def discover_gps_files(folder: Path) -> tuple[Path, Path]:
    o_env, d_env = os.environ.get("OATSIDE_ORIGIN"), os.environ.get("OATSIDE_DEST")
    if o_env and d_env:
        return Path(o_env), Path(d_env)
    cands = sorted(
        [p for p in folder.glob("Y.K._*.xlsx") if "~$" not in p.name],
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    origins = [p for p in cands if "Oatside" in p.name and "Oatside_PG" not in p.name]
    dests = [p for p in cands if ("P&G" in p.name or "เวลล์โกล" in p.name) and "Oatside_PG" not in p.name]
    if not origins or not dests:
        raise FileNotFoundError(
            "Need two GPS exports in Oatside/: one name containing 'Oatside', one 'P&G' or 'เวลล์โกล'."
        )
    return origins[0], dests[0]


# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------

def plate_from_label(s: str | None) -> str | None:
    if not s or not isinstance(s, str):
        return None
    m = PLATE_HEAD.match(s.strip())
    return m.group(1) if m else None


def _parse_dt(val) -> datetime | None:
    if isinstance(val, datetime):
        return val
    if isinstance(val, date) and not isinstance(val, datetime):
        return datetime.combine(val, datetime.min.time())
    if isinstance(val, str):
        s = val.strip()
        if len(s) >= 19:
            try:
                return datetime.strptime(s[:19], "%Y-%m-%d %H:%M:%S")
            except ValueError:
                pass
        try:
            return datetime.strptime(s[:10], "%Y-%m-%d")
        except ValueError:
            return None
    return None


def parse_legs(path: Path) -> list[Leg]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = None
    for name in wb.sheetnames:
        w = wb[name]
        if w.max_row >= 1 and w.cell(1, 4).value == "เวลาเข้า":
            ws = w
            break
    if ws is None:
        wb.close()
        raise ValueError(f"No equipment sheet in {path}")
    legs: list[Leg] = []
    current_plate: str | None = None
    for r in range(2, ws.max_row + 1):
        a, b, c, tin, tout = (
            ws.cell(r, 1).value, ws.cell(r, 2).value, ws.cell(r, 3).value,
            ws.cell(r, 4).value, ws.cell(r, 5).value,
        )
        c_s = str(c).strip() if c is not None else ""
        if c_s == "-----" or "-----" in c_s:
            current_plate = plate_from_label(str(b) if b else "")
            continue
        if not a or not DETAIL_KEY.match(str(a).strip()):
            continue
        tin = _parse_dt(tin)
        tout = _parse_dt(tout)
        if not tin or not tout:
            continue
        dev = str(c).strip() if c else ""
        p = plate_from_label(dev) or current_plate
        if not p:
            continue
        legs.append(Leg(row_no=str(a).strip(), plate=p, device=dev, t_in=tin, t_out=tout))
    wb.close()
    legs.sort(key=lambda x: (x.plate, x.t_out))
    return legs


# ---------------------------------------------------------------------------
# Matching
# ---------------------------------------------------------------------------

def hours(a: datetime, b: datetime) -> float:
    return (b - a).total_seconds() / 3600.0


def build_leg_timeline_by_plate(o_legs: list[Leg], d_legs: list[Leg]) -> dict[str, list[Leg]]:
    """Merge Origin+Dest legs per plate, sorted by In time (gap to next In)."""
    by: dict[str, list[Leg]] = defaultdict(list)
    for L in o_legs + d_legs:
        by[L.plate].append(L)
    for p in by:
        by[p].sort(key=lambda z: z.t_in)
    return by


def um_leg_dwell_gap_h(leg: Leg, timeline: list[Leg] | None) -> tuple[float, float | None]:
    """Dwell at stop (Out−In); gap hours from this Out to next leg In on same plate."""
    dwell = max(0.0, hours(leg.t_in, leg.t_out))
    if not timeline:
        return dwell, None
    idx = next((i for i, L in enumerate(timeline) if L is leg), None)
    if idx is None or idx + 1 >= len(timeline):
        return dwell, None
    gap = hours(leg.t_out, timeline[idx + 1].t_in)
    return dwell, gap


def customer_idle_clip_dest_wait_h(trip: Trip, cfg: OatsideConfig) -> float:
    """Subtract hours of (Dest_In, Dest_Out) overlapping customer_idle_windows for this plate."""
    raw = hours(trip.d_in, trip.d_out)
    sub = 0.0
    for w in cfg.customer_idle_windows:
        if w.plate != trip.plate:
            continue
        sub += w.overlap_hours(trip.d_in, trip.d_out)
    return max(0.0, raw - sub)


def origin24h_windows_for_plate(sorted_trips: list[Trip]) -> list[tuple[datetime, datetime, list[Trip]]]:
    """Rolling windows: each window [anchor, anchor+24h) collects all trips with Origin_In in range; next anchor = first trip not yet in any window."""
    out: list[tuple[datetime, datetime, list[Trip]]] = []
    if not sorted_trips:
        return out
    i = 0
    trs = sorted_trips
    while i < len(trs):
        anchor = trs[i].o_in
        end = anchor + timedelta(hours=24)
        bucket: list[Trip] = []
        j = i
        while j < len(trs) and trs[j].o_in < end:
            bucket.append(trs[j])
            j += 1
        out.append((anchor, end, bucket))
        i = j if j > i else i + 1
    return out


def feasible(o: Leg, d: Leg, max_travel_h: float) -> bool:
    if d.t_in < o.t_out:
        return False
    return hours(o.t_out, d.t_in) <= max_travel_h


# Max hours between end of accumulated origin leg and start of next origin leg for
# "double origin" chain-merge. If the driver clearly left the hub (gap > this), the
# next origin row is a new visit — do not merge (avoids stretching Origin_Out to a
# later shift when Dest_In is still far in the future).


def match_plate(
    origins: list[Leg], dests: list[Leg], max_travel_h: float
) -> tuple[list[tuple[Leg, Leg]], list[Leg], list[Leg]]:
    """Pair each destination (time order) with the most-recently-arrived unused origin
    (latest t_in before Dest_In) where Dest_In >= Origin_Out and travel <= max_travel_h.

    Sort origins by t_in descending: most recent hub arrival is the natural dispatch
    candidate — earlier visits have lower priority even if their t_out is later."""
    dests_sorted = sorted(dests, key=lambda x: (x.t_in, x.t_out))
    used_o: set[int] = set()
    pairs: list[tuple[Leg, Leg]] = []
    for d in dests_sorted:
        best_o: Leg | None = None
        # latest t_in first — most recent hub arrival dispatched preferentially
        for o in sorted(origins, key=lambda x: x.t_in, reverse=True):
            if id(o) in used_o:
                continue
            if not feasible(o, d, max_travel_h):
                continue
            best_o = o
            break
        if best_o is not None:
            pairs.append((best_o, d))
            used_o.add(id(best_o))
    uo = [x for x in origins if id(x) not in used_o]
    used_d = {id(dl) for _, dl in pairs}
    ud = [y for y in dests if id(y) not in used_d]
    return pairs, uo, ud


def merge_chained_origin_pairs(
    pairs: list[tuple[Leg, Leg]],
    max_gap_h: float,
) -> tuple[list[tuple[Leg, Leg]], list[Leg]]:
    """Resolve 'double origin' before one delivery: greedy pairs sorted by Origin_Out.
    If the next origin checks in before the current trip's Dest_In, merge origin legs and
    pick the destination with minimum feasible travel from merged Origin_Out (orphan others).
    Guard: if gap between accumulated Origin_Out and next Origin_In exceeds max_gap_h hours,
    do not chain-merge (separate hub visit)."""
    if not pairs:
        return [], []
    pairs = sorted(pairs, key=lambda pr: (pr[0].t_out, pr[0].t_in))
    out: list[tuple[Leg, Leg]] = []
    orphan_dests: list[Leg] = []
    i = 0
    while i < len(pairs):
        o_acc, d_acc = pairs[i]
        j = i + 1
        first_extend = True
        while j < len(pairs):
            o2, d2 = pairs[j]
            if o2.t_in >= d_acc.t_in:
                break
            gap_h = hours(o_acc.t_out, o2.t_in)
            if gap_h > max_gap_h:
                break
            o_acc = Leg(
                row_no=f"{o_acc.row_no}+{o2.row_no}",
                plate=o_acc.plate,
                device=o_acc.device,
                t_in=o_acc.t_in,
                t_out=o2.t_out,
            )
            last_out = o_acc.t_out
            pool = [d_acc, d2]
            feas = [d for d in pool if d.t_in >= last_out]
            use_pool = feas if feas else pool
            if first_extend:
                row_pref = [d for d in use_pool if d.row_no == o2.row_no]
                pick = min(
                    row_pref if row_pref else use_pool,
                    key=lambda d: (hours(last_out, d.t_in), d.t_in),
                )
                first_extend = False
            else:
                pick = d_acc if d_acc in use_pool else min(
                    use_pool, key=lambda d: (hours(last_out, d.t_in), d.t_in)
                )
            for d in pool:
                if id(d) != id(pick):
                    orphan_dests.append(d)
            d_acc = pick
            j += 1
        out.append((o_acc, d_acc))
        i = j
    return out, orphan_dests


def rematch_orphan_dests_to_origins(
    orphan_dests: list[Leg],
    candidates: list[Leg],
    max_travel_h: float,
) -> tuple[list[tuple[Leg, Leg]], list[Leg]]:
    """Pair orphan destinations with unused origin legs (same plate), min travel."""
    cands = sorted(candidates, key=lambda x: (x.t_out, x.t_in))
    used_o: set[int] = set()
    new_pairs: list[tuple[Leg, Leg]] = []
    still: list[Leg] = []
    for d in sorted(orphan_dests, key=lambda x: (x.t_in, x.t_out)):
        best_o: Leg | None = None
        best_tr = 1e9
        for o in cands:
            if id(o) in used_o:
                continue
            if not feasible(o, d, max_travel_h):
                continue
            tr = hours(o.t_out, d.t_in)
            if tr < best_tr or (tr == best_tr and (best_o is None or o.t_out < best_o.t_out)):
                best_o = o
                best_tr = tr
        if best_o is not None:
            new_pairs.append((best_o, d))
            used_o.add(id(best_o))
        else:
            still.append(d)
    return new_pairs, still


def collect_origin_row_refs(ol: Leg) -> list[str]:
    s = ol.row_no.strip()
    if not s:
        return []
    return [p.strip() for p in s.split("+") if p.strip()]


def constituent_origin_legs(ol: Leg, by_row: dict[str, Leg]) -> list[Leg]:
    refs = collect_origin_row_refs(ol)
    out: list[Leg] = []
    for r in refs:
        leg = by_row.get(r)
        if leg is not None:
            out.append(leg)
    if not out:
        return [ol]
    return out


def mark_used_origin_legs(ol: Leg, by_row: dict[str, Leg], used_o: set[int]) -> None:
    for lg in constituent_origin_legs(ol, by_row):
        used_o.add(id(lg))


# ---------------------------------------------------------------------------
# Site / IQR helpers
# ---------------------------------------------------------------------------

def site_for_plate(plate: str) -> str:
    if plate in BIGC_EXACT:
        return "BigC"
    m = re.match(r"^71-(\d+)$", plate)
    if m:
        n = int(m.group(1))
        if 8000 <= n <= 8009:
            return "BigC"
    return "LCB"


def iqr_threshold(travels: list[float]) -> float:
    if len(travels) < 4:
        return 8.0
    xs = sorted(travels)
    n = len(xs)

    def q(p: float) -> float:
        idx = p * (n - 1)
        lo = int(math.floor(idx))
        hi = int(math.ceil(idx))
        if lo == hi:
            return xs[lo]
        return xs[lo] + (xs[hi] - xs[lo]) * (idx - lo)

    q1, q3 = q(0.25), q(0.75)
    iqr = q3 - q1
    return max(8.0, q3 + 1.5 * iqr)


# ---------------------------------------------------------------------------
# Chronology guard
# ---------------------------------------------------------------------------

def _trip_legs_for_unmatched(t: Trip, origin_by_row: dict[str, Leg]) -> tuple[list[Leg], Leg]:
    fake_o = Leg(row_no=t.o_row, plate=t.plate, device=t.device, t_in=t.o_in, t_out=t.o_out)
    o_segs = constituent_origin_legs(fake_o, origin_by_row)
    if not o_segs:
        o_segs = [Leg(row_no=t.o_row, plate=t.plate, device=t.device, t_in=t.o_in, t_out=t.o_out)]
    dest_leg = Leg(
        row_no=t.d_row,
        plate=t.plate,
        device=t.device,
        t_in=t.d_in,
        t_out=t.d_out,
    )
    return o_segs, dest_leg


def demote_chronology_violations(
    trips: list[Trip],
    unmatched_rows: list[tuple[str, Leg, str]],
    origin_by_row_by_plate: dict[str, dict[str, Leg]],
) -> None:
    """Per plate: sort by Origin_In. If a trip Origin_In is strictly before the previous trip's
    Dest_Out, the previous match is impossible in real sequence — demote the *previous* trip to
    Unmatched (origin segment(s) + destination) and repeat until stable."""
    by_plate: dict[str, list[Trip]] = defaultdict(list)
    for t in trips:
        by_plate[t.plate].append(t)
    rebuilt: list[Trip] = []
    for plate in sorted(by_plate.keys()):
        lst = sorted(by_plate[plate], key=lambda t: (t.o_in, t.d_in))
        br = origin_by_row_by_plate.get(plate, {})
        changed = True
        while changed and len(lst) > 1:
            changed = False
            for i in range(1, len(lst)):
                if lst[i].o_in < lst[i - 1].d_out:
                    bad = lst.pop(i - 1)
                    o_segs, dleg = _trip_legs_for_unmatched(bad, br)
                    for ol in o_segs:
                        unmatched_rows.append(("Origin", ol, plate))
                    unmatched_rows.append(("Destination", dleg, plate))
                    changed = True
                    break
        rebuilt.extend(lst)
    trips[:] = sorted(rebuilt, key=lambda t: (t.plate, t.o_in))


# ---------------------------------------------------------------------------
# Build trips
# ---------------------------------------------------------------------------

def build_trips(
    origin_path: Path, dest_path: Path, cfg: OatsideConfig
) -> tuple[list[Trip], list[tuple[str, Leg, str]], list[float]]:
    o_legs = parse_legs(origin_path)
    d_legs = parse_legs(dest_path)
    by_plate_o: dict[str, list[Leg]] = defaultdict(list)
    by_plate_d: dict[str, list[Leg]] = defaultdict(list)
    for x in o_legs:
        by_plate_o[x.plate].append(x)
    for x in d_legs:
        by_plate_d[x.plate].append(x)
    plates = sorted(set(by_plate_o) | set(by_plate_d))
    trips: list[Trip] = []
    unmatched_rows: list[tuple[str, Leg, str]] = []
    origin_by_row_by_plate: dict[str, dict[str, Leg]] = {}
    mx = cfg.max_travel_h
    for p in plates:
        all_o = by_plate_o[p]
        by_row: dict[str, Leg] = {}
        for o in all_o:
            if o.row_no not in by_row:
                by_row[o.row_no] = o
        origin_by_row_by_plate[p] = by_row
        pairs, uo, ud = match_plate(all_o, by_plate_d[p], mx)
        if cfg.enable_origin_chain_merge:
            merged_pairs, orphan_d = merge_chained_origin_pairs(pairs, cfg.max_origin_chain_gap_h)
        else:
            merged_pairs, orphan_d = pairs, []
        used_o: set[int] = set()
        for ol, _ in merged_pairs:
            mark_used_origin_legs(ol, by_row, used_o)
        candidates = [o for o in all_o if id(o) not in used_o]
        rematch_pairs, still_orphan = rematch_orphan_dests_to_origins(orphan_d, candidates, mx)
        for ol, _ in rematch_pairs:
            mark_used_origin_legs(ol, by_row, used_o)
        pairs_final = merged_pairs + rematch_pairs
        for ol, dl in pairs_final:
            segs = constituent_origin_legs(ol, by_row)
            ow = sum(hours(x.t_in, x.t_out) for x in segs) if len(segs) > 1 else hours(ol.t_in, ol.t_out)
            tr = hours(ol.t_out, dl.t_in)
            dw = hours(dl.t_in, dl.t_out)
            tc = hours(segs[0].t_in, dl.t_out)
            trips.append(
                Trip(
                    plate=p,
                    site=site_for_plate(p),
                    device=segs[0].device,
                    o_row=ol.row_no,
                    d_row=dl.row_no,
                    o_in=segs[0].t_in,
                    o_out=ol.t_out,
                    d_in=dl.t_in,
                    d_out=dl.t_out,
                    origin_wait_h=ow,
                    travel_h=tr,
                    dest_wait_h=dw,
                    total_cycle_h=tc,
                    origin_date=segs[0].t_in.date(),
                    dest_date=dl.t_in.date(),
                    trip_date=segs[0].t_in.date(),
                    travel_flag=None,
                )
            )
        for ol in (o for o in all_o if id(o) not in used_o):
            unmatched_rows.append(("Origin", ol, p))
        for dl in ud + still_orphan:
            unmatched_rows.append(("Destination", dl, p))
    demote_chronology_violations(trips, unmatched_rows, origin_by_row_by_plate)
    travels = [t.travel_h for t in trips]
    thr = iqr_threshold(travels)
    for t in trips:
        t.travel_flag = "ABNORMAL" if t.travel_h >= thr else None
    return trips, unmatched_rows, travels


# ---------------------------------------------------------------------------
# Billing calculations
# ---------------------------------------------------------------------------

def base_trips_revenue_baht(trips: list[Trip], cfg: OatsideConfig) -> int:
    """Sum per-trip rate by Dest_In calendar day."""
    return sum(trip_rate_baht(t.dest_date, cfg) for t in trips)


def one_trip_fifty_pct_details_origin24h(
    trips: list[Trip],
    overrides: dict[tuple[str, date], dict[str, Any]],
    cfg: OatsideConfig,
) -> tuple[list[dict], int]:
    """+50% of one trip rate when a rolling 24h window from Origin_In contains exactly 1 matched trip. At most one origin24h surcharge per (plate, dest_date) (avoid double 24h windows on same Dest_In calendar day)."""
    by_pl: dict[str, list[Trip]] = defaultdict(list)
    for t in trips:
        by_pl[t.plate].append(t)
    rows: list[dict] = []
    total = 0
    origin24h_charged_dest: set[tuple[str, date]] = set()
    for plate in sorted(by_pl.keys()):
        lst = sorted(by_pl[plate], key=lambda x: x.o_in)
        for anchor, end, bucket in origin24h_windows_for_plate(lst):
            n = len(bucket)
            if n != 1:
                continue
            t0 = bucket[0]
            d = t0.dest_date
            key = (plate, d)
            if key in origin24h_charged_dest:
                continue
            origin24h_charged_dest.add(key)
            ov = overrides.get(key, {})
            action = ov.get("action", "")
            note = ov.get("note", "")
            if action == "exclude_50":
                continue
            if action == "include_50":
                pass
            rate = trip_rate_baht(d, cfg)
            sur = int(round(rate * cfg.one_trip_surcharge_pct / 100))
            rows.append(
                {
                    "dest_date": d,
                    "window_anchor": anchor,
                    "window_end": end,
                    "plate": plate,
                    "site": site_for_plate(plate),
                    "trips_that_day": n,
                    "auto_1trip": True,
                    "override_action": action,
                    "override_note": note,
                    "trip_rate_baht": rate,
                    "surcharge_baht": sur,
                    "fifty_kind": "origin24h",
                }
            )
            total += sur
    return rows, total


def one_trip_fifty_pct_details(
    trips: list[Trip],
    overrides: dict[tuple[str, date], dict[str, Any]],
    cfg: OatsideConfig,
) -> tuple[list[dict], int]:
    """50% surcharge on a Dest_In calendar day when exactly 1 matched trip (unless overridden)."""
    by_pd: dict[tuple[str, date], list[Trip]] = defaultdict(list)
    for t in trips:
        by_pd[(t.plate, t.dest_date)].append(t)

    rows: list[dict] = []
    total = 0
    for (plate, d), lst in sorted(by_pd.items(), key=lambda x: (x[0][1], x[0][0])):
        key = (plate, d)
        ov = overrides.get(key, {})
        action = ov.get("action", "")
        note = ov.get("note", "")
        n = len(lst)
        auto_apply = n == 1
        if action == "exclude_50":
            apply_charge = False
        elif action == "include_50":
            apply_charge = True
        else:
            apply_charge = auto_apply
        if not apply_charge:
            continue
        rate = trip_rate_baht(d, cfg)
        sur = int(round(rate * cfg.one_trip_surcharge_pct / 100))
        rows.append(
            {
                "dest_date": d,
                "plate": plate,
                "site": site_for_plate(plate),
                "trips_that_day": n,
                "auto_1trip": auto_apply,
                "override_action": action or "",
                "override_note": note,
                "window_anchor": "",
                "window_end": "",
                "trip_rate_baht": rate,
                "surcharge_baht": sur,
                "fifty_kind": "downtime_dest",
            }
        )
        total += sur
    return rows, total


def plate_dest_day_rows(
    trips: list[Trip],
    fifty_rows: list[dict],
    cfg: OatsideConfig,
    nw_rows: list[dict] | None = None,
) -> list[dict]:
    """Per (plate, Dest_In date): base line + sum surcharges; HTML cell can show multiple badges."""
    by_pd: dict[tuple[str, date], list[Trip]] = defaultdict(list)
    for t in trips:
        by_pd[(t.plate, t.dest_date)].append(t)
    fifty_lists: dict[tuple[str, date], list[dict]] = defaultdict(list)
    for r in fifty_rows:
        p = r.get("plate")
        d = r.get("dest_date")
        if p and isinstance(d, date):
            fifty_lists[(str(p), d)].append(r)
    nw_by: dict[tuple[str, date], dict] = {}
    if nw_rows:
        for nr in nw_rows:
            nw_by[(str(nr["plate"]), nr["dest_date"])] = nr
    out: list[dict] = []
    seen_keys: set[tuple[str, date]] = set()
    for (plate, d), lst in sorted(by_pd.items(), key=lambda x: (x[0][1], x[0][0])):
        key = (str(plate), d)
        seen_keys.add(key)
        rate = trip_rate_baht(d, cfg)
        n = len(lst)
        base_line = n * rate
        frs = fifty_lists.get(key, [])
        sur = sum(int(x.get("surcharge_baht", 0) or 0) for x in frs)
        badge_parts: list[str] = []
        for x in frs:
            b = html_fifty_surcharge_badge(x, cfg)
            if b:
                badge_parts.append(b)
        nr = nw_by.get(key)
        if nr:
            ns = int(nr.get("surcharge_baht", 0) or 0)
            if ns > 0:
                sur += ns
                synth = {
                    "plate": plate,
                    "dest_date": d,
                    "trip_rate_baht": int(nr.get("trip_rate_baht", 0) or 0),
                    "surcharge_baht": ns,
                    "fifty_kind": "no_work_outbound",
                }
                b2 = html_fifty_surcharge_badge(synth, cfg)
                if b2:
                    badge_parts.append(b2)
        badge = " ".join(badge_parts) if badge_parts else ""
        out.append(
            {
                "dest_date": d,
                "plate": plate,
                "site": site_for_plate(plate),
                "matched_trips": n,
                "trip_rate_baht": rate,
                "base_line_baht": base_line,
                "fifty_pct_baht": sur,
                "fifty_badge_html": badge,
                "return_trip_baht": 0,
                "customer_day_baht": base_line + sur,
            }
        )

    # Synthetic rows: recovery No-work anchor day may have no matched Dest_In that calendar date
    if nw_rows:
        for nr in nw_rows:
            nk = (str(nr["plate"]), nr["dest_date"])
            if nk in seen_keys:
                continue
            seen_keys.add(nk)
            plate, d = nk[0], nk[1]
            rate = trip_rate_baht(d, cfg)
            frs = fifty_lists.get(nk, [])
            sur = sum(int(x.get("surcharge_baht", 0) or 0) for x in frs)
            badge_parts: list[str] = []
            for x in frs:
                b = html_fifty_surcharge_badge(x, cfg)
                if b:
                    badge_parts.append(b)
            ns = int(nr.get("surcharge_baht", 0) or 0)
            if ns > 0:
                sur += ns
                synth = {
                    "plate": plate,
                    "dest_date": d,
                    "trip_rate_baht": int(nr.get("trip_rate_baht", 0) or 0),
                    "surcharge_baht": ns,
                    "fifty_kind": "no_work_outbound",
                }
                b2 = html_fifty_surcharge_badge(synth, cfg)
                if b2:
                    badge_parts.append(b2)
            badge = " ".join(badge_parts) if badge_parts else ""
            out.append(
                {
                    "dest_date": d,
                    "plate": plate,
                    "site": site_for_plate(plate),
                    "matched_trips": 0,
                    "trip_rate_baht": rate,
                    "base_line_baht": 0,
                    "fifty_pct_baht": sur,
                    "fifty_badge_html": badge,
                    "customer_day_baht": sur,
                }
            )

    out.sort(key=lambda r: (r["dest_date"], str(r["plate"])))
    return out


    return out


def customer_trips_per_day_rows(trips: list[Trip]) -> list[dict]:
    """Matched trips aggregated by Dest_In calendar day (fleet total) for customer one-pager."""
    by_t: dict[date, int] = defaultdict(int)
    by_plates: dict[date, set[str]] = defaultdict(set)
    for tr in trips:
        d = tr.dest_date
        by_t[d] += 1
        by_plates[d].add(tr.plate)
    return [
        {"dest_date": d, "matched_trips": by_t[d], "active_trucks": len(by_plates[d])}
        for d in sorted(by_t.keys())
    ]


def audit_log_rows(

    trips: list[Trip],
    fifty_rows: list[dict],
    overrides: dict[tuple[str, date], dict[str, Any]],
    cfg: OatsideConfig,
) -> list[dict]:
    """Per (plate, dest_date): one row explaining the billing decision in Thai."""
    by_pd: dict[tuple[str, date], list[Trip]] = defaultdict(list)
    for t in trips:
        by_pd[(t.plate, t.dest_date)].append(t)
    fifty_key = {(r["plate"], r["dest_date"]): r for r in fifty_rows}

    rows: list[dict] = []
    for (plate, d), lst in sorted(by_pd.items(), key=lambda x: (x[0][1], x[0][0])):
        n = len(lst)
        rate = trip_rate_baht(d, cfg)
        base = n * rate
        fr = fifty_key.get((plate, d))
        sur = int(fr["surcharge_baht"]) if fr else 0
        total = base + sur

        ov = overrides.get((plate, d), {})
        action = ov.get("action", "")
        note = ov.get("note", "")

        if action == "exclude_50":
            fifty_rule = f"ไม่เก็บ {cfg.one_trip_surcharge_pct:.0f}% [override: exclude_50]"
            if note:
                fifty_rule += f" — {note}"
        elif action == "include_50":
            fifty_rule = f"เก็บ {cfg.one_trip_surcharge_pct:.0f}% [override: include_50]"
            if note:
                fifty_rule += f" — {note}"
        elif n == 1:
            fifty_rule = f"เก็บ {cfg.one_trip_surcharge_pct:.0f}% อัตโนมัติ (วิ่ง 1 เที่ยว)"
        else:
            fifty_rule = f"ไม่เก็บ {cfg.one_trip_surcharge_pct:.0f}% (วิ่ง {n} เที่ยว)"

        rows.append({
            "dest_date": d,
            "plate": plate,
            "site": site_for_plate(plate),
            "matched_trips": n,
            "trip_rate_baht": rate,
            "base_line_baht": base,
            "fifty_pct_baht": sur,
            "customer_day_baht": total,
            "billing_note": fifty_rule,
        })
    return rows


def one_trip_fifty_pct_origin_day(
    trips: list[Trip],
    overrides: dict[tuple[str, date], dict[str, Any]],
    cfg: OatsideConfig,
) -> tuple[list[dict], int]:
    """50% surcharge when exactly 1 matched trip on Origin_In calendar day (วันงาน).
    Rules (Session #94): วันงาน = Origin_In date; ข้ามคืนไม่แตกวัน; วันไม่มี Origin ไม่นับ.
    Override key uses dest_date for compatibility with oatside_billing_overrides.json."""
    by_pd: dict[tuple[str, date], list[Trip]] = defaultdict(list)
    for t in trips:
        by_pd[(t.plate, t.o_in.date())].append(t)
    rows: list[dict] = []
    total = 0
    for (plate, origin_day), lst in sorted(by_pd.items(), key=lambda x: (x[0][1], x[0][0])):
        n = len(lst)
        if n != 1:
            continue
        t0 = lst[0]
        key = (plate, t0.dest_date)
        ov = overrides.get(key, {})
        action = ov.get("action", "")
        note = ov.get("note", "")
        if action == "exclude_50":
            continue
        rate = trip_rate_baht(t0.dest_date, cfg)
        sur = int(round(rate * cfg.one_trip_surcharge_pct / 100))
        note_l = (note or "").lower()
        fifty_kind = (
            "blank_run"
            if (action == "blank_run" or "ตีเปล่า" in note_l)
            else "downtime_origin_day"
        )
        rows.append({
            "origin_day": origin_day,
            "dest_date": t0.dest_date,
            "plate": plate,
            "site": site_for_plate(plate),
            "trips_that_day": n,
            "auto_1trip": True,
            "override_action": action,
            "override_note": note,
            "window_anchor": str(origin_day),
            "window_end": "",
            "trip_rate_baht": rate,
            "surcharge_baht": sur,
            "fifty_kind": fifty_kind,
        })
        total += sur
    return rows, total




def supplement_long_dest_wait_midnight_fifty(
    trips: list[Trip],
    fifty_rows: list[dict],
    overrides: dict[tuple[str, date], dict[str, Any]],
    cfg: OatsideConfig,
) -> tuple[list[dict], int]:
    """Dest_In->Dest_Out crosses midnight, dwell >= min_h: add surcharge keyed by (plate, dest_date)
    when no fifty row yet. Default: full 1-trip rate (not 50pct) — idle calendar day at customer."""
    if not getattr(cfg, "long_dest_wait_midnight_fifty", True):
        return [], 0
    min_h = float(getattr(cfg, "long_dest_wait_midnight_min_h", 12.0))
    full_trip = bool(getattr(cfg, "long_dest_wait_midnight_full_trip", True))
    charged: dict[tuple[str, date], int] = {}
    for r in fifty_rows:
        p = r.get("plate")
        d = r.get("dest_date")
        if p and isinstance(d, date):
            charged[(str(p), d)] = int(r.get("surcharge_baht", 0) or 0)
    extra: list[dict] = []
    total = 0
    for t in trips:
        if t.d_in.date() >= t.d_out.date():
            continue
        if t.dest_wait_h < min_h:
            continue
        key = (t.plate, t.dest_date)
        if charged.get(key, 0) > 0:
            continue
        ov = overrides.get(key, {})
        if ov.get("action") == "exclude_50":
            continue
        rate = trip_rate_baht(t.dest_date, cfg)
        if full_trip:
            sur = int(rate)
            pct_note = "เต็ม 1 เที่ยว (เรทวัน Dest_In)"
        else:
            sur = int(round(rate * float(cfg.one_trip_surcharge_pct) / 100.0))
            pct_note = f"+{cfg.one_trip_surcharge_pct:.0f}% เรทวัน Dest_In"
        note = (
            f"รอปลายทางข้ามคืน Dest_In→Dest_Out ({t.dest_wait_h:.2f}h); {pct_note}"
        )
        extra.append(
            {
                "origin_day": t.o_in.date(),
                "dest_date": t.dest_date,
                "plate": t.plate,
                "site": site_for_plate(t.plate),
                "trips_that_day": 1,
                "auto_1trip": False,
                "override_action": ov.get("action", "") or "",
                "override_note": (ov.get("note", "") or "") + ("; " if ov.get("note") else "") + note,
                "window_anchor": str(t.d_in),
                "window_end": str(t.d_out),
                "trip_rate_baht": rate,
                "surcharge_baht": sur,
                "fifty_kind": ("midnight_full" if full_trip else "midnight_pct"),
            }
        )
        charged[key] = sur
        total += sur
    return extra, total


def origin_day_audit_rows(
    trips: list[Trip],
    fifty_rows: list[dict],
    overrides: dict[tuple[str, date], dict[str, Any]],
    cfg: OatsideConfig,
) -> list[dict]:
    """Per (plate, origin_day): billing explanation grouped by Origin_In calendar date."""
    by_pd: dict[tuple[str, date], list[Trip]] = defaultdict(list)
    for t in trips:
        by_pd[(t.plate, t.o_in.date())].append(t)
    fifty_key = {(r["plate"], r["origin_day"]): r for r in fifty_rows if "origin_day" in r}
    rows: list[dict] = []
    for (plate, origin_day), lst in sorted(by_pd.items(), key=lambda x: (x[0][1], x[0][0])):
        n = len(lst)
        t0 = min(lst, key=lambda x: x.o_in)
        rate = trip_rate_baht(t0.dest_date, cfg)
        base = n * rate
        fr = fifty_key.get((plate, origin_day))
        sur = int(fr["surcharge_baht"]) if fr else 0
        total = base + sur
        ov = overrides.get((plate, t0.dest_date), {})
        action = ov.get("action", "")
        note = ov.get("note", "")
        if action == "exclude_50":
            fifty_rule = f"ไม่เก็บ {cfg.one_trip_surcharge_pct:.0f}% [override: exclude_50]"
            if note:
                fifty_rule += f" — {note}"
        elif action == "include_50":
            fifty_rule = f"เก็บ {cfg.one_trip_surcharge_pct:.0f}% [override: include_50]"
            if note:
                fifty_rule += f" — {note}"
        elif n == 1:
            fifty_rule = f"เก็บ {cfg.one_trip_surcharge_pct:.0f}% อัตโนมัติ (วันงาน 1 เที่ยว)"
        else:
            fifty_rule = f"ไม่เก็บ {cfg.one_trip_surcharge_pct:.0f}% (วันงาน {n} เที่ยว)"
        rows.append({
            "origin_day": origin_day,
            "dest_date": t0.dest_date,
            "plate": plate,
            "site": site_for_plate(plate),
            "matched_trips": n,
            "trip_rate_baht": rate,
            "base_line_baht": base,
            "fifty_pct_baht": sur,
            "customer_day_baht": total,
            "billing_note": fifty_rule,
        })
    return rows


def daily_activity_by_dest(trips: Iterable[Trip], cfg: OatsideConfig) -> list[tuple[date, dict]]:
    """Return sorted list of (dest_date, stats)."""
    by: dict[date, dict] = defaultdict(
        lambda: {
            "plates": set(),
            "trips": 0,
            "bigc_p": set(),
            "bigc_t": 0,
            "lcb_p": set(),
            "lcb_t": 0,
        }
    )
    for t in trips:
        d = t.dest_date
        by[d]["plates"].add(t.plate)
        by[d]["trips"] += 1
        if t.site == "BigC":
            by[d]["bigc_p"].add(t.plate)
            by[d]["bigc_t"] += 1
        else:
            by[d]["lcb_p"].add(t.plate)
            by[d]["lcb_t"] += 1
    out = []
    for d in sorted(by):
        s = by[d]
        trucks = len(s["plates"])
        commit = cfg.min_trips_per_truck * trucks
        short = max(0, commit - s["trips"])
        out.append((d, {**s, "trucks": trucks, "commit": commit, "short": short}))
    return out


def billing_totals(rows: list[tuple[date, dict]], cfg: OatsideConfig) -> tuple[int, int, int, int]:
    actual = sum(s["trips"] for _, s in rows)
    commit = sum(s["commit"] for _, s in rows)
    short = sum(s["short"] for _, s in rows)
    extra = 0
    for d, s in rows:
        r = trip_rate_baht(d, cfg)
        extra += s["short"] * r
    return actual, commit, short, extra


def site_billing(rows: list[tuple[date, dict]], cfg: OatsideConfig) -> tuple[int, int, int, int, int, int, int, int]:
    """Returns BigC (actual, commit, short, extra) then LCB (actual, commit, short, extra)."""
    bc_a = bc_c = bc_s = bc_e = 0
    lc_a = lc_c = lc_s = lc_e = 0
    for d, s in rows:
        r = trip_rate_baht(d, cfg)
        bt = s["bigc_t"]
        bc_min = cfg.min_trips_per_truck * len(s["bigc_p"])
        bs = max(0, bc_min - bt)
        lt = s["lcb_t"]
        lc_min = cfg.min_trips_per_truck * len(s["lcb_p"])
        ls = max(0, lc_min - lt)
        bc_a += bt
        bc_c += bc_min
        bc_s += bs
        bc_e += bs * r
        lc_a += lt
        lc_c += lc_min
        lc_s += ls
        lc_e += ls * r
    return bc_a, bc_c, bc_s, bc_e, lc_a, lc_c, lc_s, lc_e


def daily_time_rows(
    trips: list[Trip], unmatched: list[tuple[str, Leg, str]], cfg: OatsideConfig
) -> list[tuple]:
    matched_cycle_h: dict[tuple[date, str], float] = defaultdict(float)
    matched_origin_wait_h: dict[tuple[date, str], float] = defaultdict(float)
    matched_dest_wait_h: dict[tuple[date, str], float] = defaultdict(float)
    matched_travel_h: dict[tuple[date, str], float] = defaultdict(float)
    for t in trips:
        key = (t.trip_date, t.plate)
        dw_raw = t.dest_wait_h
        dw_adj = customer_idle_clip_dest_wait_h(t, cfg)
        cycle_adj = t.total_cycle_h - max(0.0, dw_raw - dw_adj)
        matched_cycle_h[key] += max(0.0, cycle_adj)
        matched_origin_wait_h[key] += t.origin_wait_h
        matched_dest_wait_h[key] += dw_adj
        matched_travel_h[key] += t.travel_h
    uo: dict[tuple[date, str], float] = defaultdict(float)
    ud: dict[tuple[date, str], float] = defaultdict(float)
    for src, leg, _p in unmatched:
        key = (leg.t_in.date(), leg.plate)
        h = hours(leg.t_in, leg.t_out)
        if h < 0 or h > 72:
            continue
        if src == "Origin":
            uo[key] += h
        else:
            h2 = h
            for w in cfg.customer_idle_windows:
                if w.plate == leg.plate:
                    h2 -= w.overlap_hours(leg.t_in, leg.t_out)
            ud[key] += max(0.0, h2)
    keys = sorted(
        set(matched_cycle_h) | set(matched_origin_wait_h) | set(matched_dest_wait_h)
        | set(matched_travel_h) | set(uo) | set(ud),
        key=lambda x: (x[0], x[1]),
    )
    rows = []
    for d, plate in keys:
        key = (d, plate)
        cycle_h = matched_cycle_h.get(key, 0.0)
        matched_ow = matched_origin_wait_h.get(key, 0.0)
        matched_dw = matched_dest_wait_h.get(key, 0.0)
        matched_tr = matched_travel_h.get(key, 0.0)
        um_ow = uo.get(key, 0.0)
        um_dw = ud.get(key, 0.0)
        adjusted_ow = matched_ow + um_ow
        adjusted_dw = matched_dw + um_dw
        combined_h = adjusted_ow + matched_tr + adjusted_dw
        rows.append((
            d, plate, site_for_plate(plate),
            cycle_h, matched_ow, matched_dw, matched_tr,
            um_ow, um_dw, adjusted_ow, adjusted_dw, combined_h,
            24.0 - combined_h,
        ))
    return rows



def _parse_no_work_entries(raw: object) -> list[tuple[date, date, str]]:
    out: list[tuple[date, date, str]] = []
    if not isinstance(raw, list):
        return out
    for item in raw:
        if not isinstance(item, dict):
            continue
        a = _parse_dt(item.get("from") or item.get("start"))
        b = _parse_dt(item.get("to") or item.get("end"))
        if not a or not b:
            continue
        da, db = a.date(), b.date()
        if db < da:
            da, db = db, da
        note = str(item.get("note", "")).strip()
        out.append((da, db, note))
    return out


def _parse_date_set(raw: object) -> frozenset[date]:
    if not isinstance(raw, list) or not raw:
        return frozenset()
    s: set[date] = set()
    for x in raw:
        if isinstance(x, str) and len(x) >= 10:
            try:
                s.add(datetime.strptime(x[:10], "%Y-%m-%d").date())
            except ValueError:
                continue
    return frozenset(s)


def first_matched_trip_by_plate_dest(trips: list[Trip]) -> dict[tuple[str, date], Trip]:
    by: dict[tuple[str, date], list[Trip]] = defaultdict(list)
    for t in trips:
        by[(t.plate, t.dest_date)].append(t)
    return {k: min(lst, key=lambda x: x.d_in) for k, lst in by.items()}



def first_no_work_trip_by_plate_recovery_day(
    trips: list[Trip], cfg: OatsideConfig
) -> dict[tuple[str, date], Trip]:
    """(plate, recovery_R) -> trip that carries No-work outbound +50%%.

    ``recovery_R`` is a calendar date in ``outbound_half_dest_dates`` (day after no-work block).

    1) Prefer matched trips with ``dest_date == R`` (earliest ``d_in``).
    2) Else matched trips with ``origin_date == R`` and ``dest_date > R`` (overnight; earliest ``o_in``).
       Fixes trucks that start on recovery morning but ``Dest_In`` falls next calendar day.
    """
    out: dict[tuple[str, date], Trip] = {}
    for R in cfg.outbound_half_dest_dates:
        plates = {t.plate for t in trips}
        for plate in plates:
            same_dest = [t for t in trips if t.plate == plate and t.dest_date == R]
            if same_dest:
                out[(plate, R)] = min(same_dest, key=lambda x: x.d_in)
                continue
            cross = [t for t in trips if t.plate == plate and t.origin_date == R and t.dest_date > R]
            if cross:
                out[(plate, R)] = min(cross, key=lambda x: x.o_in)
    return out


def _split_fifty_surcharge_50_100(frs: list[dict]) -> tuple[int, int]:
    """Sum fifty surcharges into +50%% bucket vs +100%% bucket (exclude no-work/blank_run rows)."""
    a50 = 0
    a100 = 0
    for r in frs:
        sur = int(r.get("surcharge_baht", 0) or 0)
        if sur <= 0:
            continue
        k = str(r.get("fifty_kind") or "")
        if k == "midnight_full":
            a100 += sur
        elif k in ("no_work_outbound", "blank_run"):
            continue
        elif k == "midnight_pct":
            a50 += sur
        elif k in ("downtime_dest", "downtime_origin_day", "origin24h"):
            a50 += sur
        else:
            rate = int(r.get("trip_rate_baht", 0) or 0)
            if rate > 0 and sur >= rate:
                a100 += sur
            else:
                a50 += sur
    return a50, a100


def trip_row_pricing_cells(
    t: Trip,
    *,
    firsts: dict[tuple[str, date], Trip],
    first_no_work: dict[tuple[str, date], Trip],
    fifty_by_lists: dict[tuple[str, date], list[dict]],
    cfg: OatsideConfig,
    return_baht: int = 0,
) -> str:
    """HTML <td>…×4 after wait columns: base rate, downtime+50, downtime+100, blank(no-work)+50."""
    rate = trip_rate_baht(t.dest_date, cfg)
    ft = firsts.get((t.plate, t.dest_date))
    frs = fifty_by_lists.get((str(t.plate), t.dest_date), [])
    dw50 = dw100 = 0
    if ft is not None and id(ft) == id(t):
        dw50, dw100 = _split_fifty_surcharge_50_100(frs)
    nw_amt = trip_no_work_outbound_baht(t, first_no_work, cfg)

    def money_td(n: int) -> str:
        return f"<td class='money'>{n:,}</td>" if n else "<td>—</td>"

    return (
        f"<td class='money'>{rate:,}</td>"
        + money_td(dw50)
        + money_td(dw100)
        + money_td(nw_amt)
        + money_td(return_baht)
    )


def no_work_outbound_rows(trips: list[Trip], cfg: OatsideConfig) -> tuple[list[dict], int]:
    """+50pct of trip rate on first matched trip after recovery calendar day R (see first_no_work_trip_by_plate_recovery_day)."""
    first_no_work = first_no_work_trip_by_plate_recovery_day(trips, cfg)
    rows: list[dict] = []
    total = 0
    pct = float(cfg.one_trip_surcharge_pct)
    for (plate, R), t0 in sorted(first_no_work.items(), key=lambda x: (x[0][1], x[0][0])):
        rate = trip_rate_baht(R, cfg)
        sur = int(round(rate * pct / 100.0))
        rows.append(
            {
                "dest_date": R,
                "plate": plate,
                "site": site_for_plate(plate),
                "d_row": t0.d_row,
                "trip_rate_baht": rate,
                "surcharge_baht": sur,
                "note": (
                    "No-work recovery: anchor "
                    f"{R} (Dest_In of chosen trip {t0.dest_date})"
                ),
            }
        )
        total += sur
    return rows, total


def phantom_zero_trip_candidates(origin_legs: list[Leg], trips: list[Trip], cfg: OatsideConfig) -> list[dict]:
    """Days with Origin legs but no matched trip on that trip_date (suggest 1 full trip charge)."""
    matched_origin_days = {(t.plate, t.trip_date) for t in trips}
    hours_by: dict[tuple[str, date], float] = defaultdict(float)
    for leg in origin_legs:
        d = leg.t_in.date()
        hours_by[(leg.plate, d)] += hours(leg.t_in, leg.t_out)
    rows: list[dict] = []
    for (plate, d), h in sorted(hours_by.items(), key=lambda x: (x[0][1], x[0][0])):
        if (plate, d) in matched_origin_days:
            continue
        if h < 1.0:
            continue
        rate = trip_rate_baht(d, cfg)
        rows.append(
            {
                "plate": plate,
                "calendar_date": d,
                "origin_hours_on_day": round(h, 2),
                "suggest_full_trip_baht": rate,
                "note": "No matched trip on this trip_date; OAT rule: charge 1 full trip (review before adding to grand total)",
            }
        )
    return rows


def double_origin_um_hints(unmatched: list[tuple[str, Leg, str]]) -> list[dict]:
    """Flag days with 2+ unmatched Origin segments (possible double hub in/out)."""
    by: dict[tuple[str, date], int] = defaultdict(int)
    for src, leg, plate in unmatched:
        if src != "Origin":
            continue
        by[(plate, leg.t_in.date())] += 1
    return [
        {
            "plate": plate,
            "calendar_date": d,
            "um_origin_segments": n,
            "note": "2+ unmatched Origin rows same calendar day — review",
        }
        for (plate, d), n in sorted(by.items(), key=lambda x: (x[0][1], x[0][0]))
        if n >= 2
    ]


def trip_no_work_outbound_baht(
    t: Trip, first_no_work: dict[tuple[str, date], Trip], cfg: OatsideConfig
) -> int:
    pct = float(cfg.one_trip_surcharge_pct)
    for R in cfg.outbound_half_dest_dates:
        ft = first_no_work.get((t.plate, R))
        if ft is None or id(ft) != id(t):
            continue
        rate = trip_rate_baht(R, cfg)
        return int(round(rate * pct / 100.0))
    return 0




# ---------------------------------------------------------------------------
# Excel styling & per-table exports (ลูกค้า)
# ---------------------------------------------------------------------------

OATSIDE_EXPORT_TABLES: list[tuple[str, str, str]] = [
    ("Customer_Trips_Per_Day", "01_CPD_MatchedTripsPerDay.xlsx", "(1) จำนวนเที่ยวต่อวัน (matched Dest_In)"),
    ("Plate_DestDay", "02_Plate_DestDay_Daily.xlsx", "(2) เดลี่รถทุกคัน — Dest_In × ทะเบียน"),
    ("Unmatched_Log", "03_Unmatched_Legs.xlsx", "(3) Unmatched legs"),
    ("Audit_Log", "04_Audit_Log.xlsx", "Audit Log — เหตุผลการคิดเงิน"),
    ("Trip_Detail", "05_Trip_Detail.xlsx", "รายเที่ยว Trip Detail"),
    ("Customer_Summary", "06_Customer_Summary.xlsx", "สรุปลูกค้า (บรรทัด A/B/C/D)"),
    ("Daily_Activity", "07_Daily_Activity.xlsx", "Daily Activity (รวมไซท์)"),
    ("Daily_Time_24h_Check", "08_Daily_Time_24h_Check.xlsx", "Daily Time 24h Check"),
    ("Surcharge_50pct_1Trip", "09_Surcharge_50pct_1Trip.xlsx", "Surcharge 50% / 100% / ตีเปล่า (รายทะเบียน×วัน)"),
    ("Manual_Extra_Trips", "10_Manual_Extra_Trips.xlsx", "เที่ยวเพิ่ม (manual_extra_trips)"),
    ("Manual_Return_Trips", "11_Manual_Return_Trips.xlsx", "ค่าขนส่งขากลับ (manual_return_trips)"),
    ("NoWork_Outbound_50pct", "12_NoWork_Outbound_50pct.xlsx", "No-work recovery outbound 50%"),
    ("Phantom_Trip_Candidates", "13_Phantom_Trip_Candidates.xlsx", "Phantom trip candidates"),
    ("Hints_DoubleOrigin", "14_Hints_DoubleOrigin.xlsx", "Hints double-origin (UM)"),
]


def _hdr_moneyish(cell_val) -> bool:
    if cell_val is None:
        return False
    s = str(cell_val).lower()
    t = str(cell_val)
    return ("฿" in t) or ("baht" in s) or ("บาท" in t)


def _thin_border():
    from openpyxl.styles import Border, Side

    t = Side(style="thin", color="CCD6E4")
    return Border(left=t, right=t, top=t, bottom=t)


def beautify_oatside_workbook(wb) -> None:
    """Apply consistent table styling to all sheets (Info = compact key/value)."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    head_fill = PatternFill("solid", fgColor="1E3A5F")
    head_font = Font(color="FFFFFF", bold=True, size=11)
    zebra = PatternFill("solid", fgColor="F4F7FB")
    title_font = Font(bold=True, size=12, color="1E3A5F")
    bdr = _thin_border()

    for name in wb.sheetnames:
        ws = wb[name]
        if ws.max_row == 0 or ws.max_column == 0:
            continue
        if name == "Info":
            for r in range(1, ws.max_row + 1):
                a = ws.cell(r, 1)
                b = ws.cell(r, 2)
                a.font = title_font if r == 1 else Font(bold=True, color="2C3E50")
                a.alignment = Alignment(vertical="top", wrap_text=True)
                if b.value is not None:
                    b.alignment = Alignment(vertical="top", wrap_text=True)
                a.border = bdr
                b.border = bdr
            ws.column_dimensions["A"].width = 34
            ws.column_dimensions["B"].width = 86
            continue

        hdr_row = 1
        last_c = ws.max_column
        last_r = ws.max_row
        money_cols: set[int] = set()
        for c in range(1, last_c + 1):
            hv = ws.cell(hdr_row, c).value
            if _hdr_moneyish(hv):
                money_cols.add(c)
        for c in range(1, last_c + 1):
            ch = get_column_letter(c)
            cell = ws.cell(hdr_row, c)
            cell.fill = head_fill
            cell.font = head_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = bdr
            maxlen = 10
            for r in range(1, last_r + 1):
                v = ws.cell(r, c).value
                if v is None:
                    continue
                s = str(v)
                maxlen = max(maxlen, min(len(s), 48))
            ws.column_dimensions[ch].width = min(52, max(10, maxlen + 2))
        for r in range(hdr_row + 1, last_r + 1):
            fill = zebra if (r % 2 == 0) else None
            for c in range(1, last_c + 1):
                cell = ws.cell(r, c)
                cell.border = bdr
                if fill is not None:
                    cell.fill = fill
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if c in money_cols and isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0"
        ws.freeze_panes = f"A{hdr_row + 1}"
        ws.auto_filter.ref = f"A{hdr_row}:{get_column_letter(last_c)}{last_r}"


def write_split_excel_exports(wb_path: Path, report_dir: Path, *, built_at: str) -> None:
    """Write one .xlsx per customer-facing table under report_dir/exports/."""
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    exp = report_dir / "exports"
    exp.mkdir(parents=True, exist_ok=True)
    src = load_workbook(wb_path, data_only=False)
    head_fill = PatternFill("solid", fgColor="1E3A5F")
    head_font = Font(color="FFFFFF", bold=True, size=11)
    zebra = PatternFill("solid", fgColor="F4F7FB")
    brand_font = Font(bold=True, size=14, color="FFFFFF")
    sub_font = Font(size=11, color="2C3E50")
    bdr = _thin_border()

    for sheet_name, fname, th_label in OATSIDE_EXPORT_TABLES:
        if sheet_name not in src.sheetnames:
            continue
        sws = src[sheet_name]
        if sws.max_row == 0:
            continue
        nb = Workbook()
        tws = nb.active
        tws.title = sheet_name[:31]
        mc = max(6, sws.max_column)
        end_l = get_column_letter(mc)
        tws.merge_cells(f"A1:{end_l}1")
        c1 = tws["A1"]
        c1.value = "Y.K. Logistics — Oatside / P&G"
        c1.font = brand_font
        c1.fill = head_fill
        c1.alignment = Alignment(horizontal="center", vertical="center")
        tws.row_dimensions[1].height = 26
        tws.append([th_label, built_at])
        tws["A2"].font = Font(bold=True, size=12, color="1E3A5F")
        tws["B2"].font = sub_font
        tws.append([""] * mc)
        hdr_r = 4
        for r in range(1, sws.max_row + 1):
            for c in range(1, sws.max_column + 1):
                tws.cell(hdr_r + r - 1, c).value = sws.cell(r, c).value
        last_r = tws.max_row
        last_c = tws.max_column
        money_cols: set[int] = set()
        for c in range(1, last_c + 1):
            if _hdr_moneyish(tws.cell(hdr_r, c).value):
                money_cols.add(c)
        for c in range(1, last_c + 1):
            ch = get_column_letter(c)
            cell = tws.cell(hdr_r, c)
            cell.fill = head_fill
            cell.font = head_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = bdr
            maxlen = 10
            for r in range(hdr_r, last_r + 1):
                v = tws.cell(r, c).value
                if v is None:
                    continue
                s = str(v)
                maxlen = max(maxlen, min(len(s), 48))
            tws.column_dimensions[ch].width = min(52, max(10, maxlen + 2))
        for r in range(hdr_r + 1, last_r + 1):
            fill = zebra if (r % 2 == 0) else None
            for c in range(1, last_c + 1):
                cell = tws.cell(r, c)
                cell.border = bdr
                if fill is not None:
                    cell.fill = fill
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if c in money_cols and isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0"
        tws.freeze_panes = f"A{hdr_r + 1}"
        tws.auto_filter.ref = f"A{hdr_r}:{get_column_letter(last_c)}{last_r}"
        nb.save(exp / fname)
        nb.close()
    src.close()


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def write_excel(
    path: Path,
    origin_name: str,
    dest_name: str,
    trips: list[Trip],
    unmatched: list[tuple[str, Leg, str]],
    daily_time: list[tuple],
    daily_rows: list[tuple[date, dict]],
    fifty_rows: list[dict],
    fifty_total_baht: int,
    min_trip_extra_baht: int,
    audit_rows: list[dict],
    cfg: OatsideConfig,
    customer_grand_baht: int,
    no_work_rows: list[dict],
    no_work_total_baht: int,
    phantom_rows: list[dict],
    hint_rows: list[dict],
    pday_rows: list[dict],
    cpd_rows: list[dict],
    leg_timeline_by_plate: dict[str, list[Leg]],
) -> None:
    base_baht = base_trips_revenue_baht(trips, cfg) + sum_manual_extra_baht(cfg)
    pday = pday_rows
    wb = openpyxl.Workbook()
    default = wb.active
    wb.remove(default)

    # --- Info ---
    info = wb.create_sheet("Info", 0)
    info.append(["Built", datetime.now().strftime("%Y-%m-%d %H:%M")])
    info.append(["Origin file", origin_name])
    info.append(["Dest file", dest_name])
    info.append(["Config file", str(_config_path())])
    info.append(["Max_travel_h", cfg.max_travel_h])
    info.append(["Max_origin_chain_gap_h", cfg.max_origin_chain_gap_h])
    info.append(["Enable_origin_chain_merge", cfg.enable_origin_chain_merge])
    info.append(["Min_trips_per_truck_per_day", cfg.min_trips_per_truck])
    info.append(["One_trip_surcharge_pct", cfg.one_trip_surcharge_pct])
    info.append(["Trip_rates", config_rate_summary(cfg)])
    info.append(["Matcher",
        f"Greedy min-travel; feasible if Dest_In>=Origin_Out and travel<={cfg.max_travel_h}h"])
    info.append(["Surcharge_50pct_1Trip",
        f"If exactly 1 matched trip on Dest_In day -> add {cfg.one_trip_surcharge_pct:.0f}% of trip rate. "
        f"Overrides: {_overrides_json_path()} (exclude_50 / include_50)"])
    info.append(["Base_trips_revenue_baht", base_baht])
    info.append(["Manual_extra_trips_baht", sum_manual_extra_baht(cfg)])
    info.append(["Manual_return_trips_baht", sum_manual_return_baht(cfg)])
    info.append(["Use_origin_24h_fifty", cfg.use_origin_24h_fifty])
    info.append(["Customer_idle_windows", len(cfg.customer_idle_windows)])
    info.append(["Charge_min_trip_shortfall", cfg.charge_min_trip_shortfall])
    info.append(["Min2trips_extra_baht", min_trip_extra_baht])
    info.append(["Fifty_pct_surcharge_total_baht", fifty_total_baht])
    info.append(["No_work_outbound_50pct_total_baht", no_work_total_baht])
    info.append(
        [
            "Policy_recovery_plus_fifty",
            "เก็บคู่: วัน recovery เที่ยวแรกอาจได้ทั้ง surcharge fifty (ดาวน์ไทม์) และ No-work outbound 50pct — บวกทั้งคู่ตามนโยบายผู้ใช้ 2026-05-01",
        ]
    )
    info.append(["Phantom_zero_trip_candidates", len(phantom_rows)])
    info.append(["Double_origin_um_hints", len(hint_rows)])
    cg_note = (
        "base + min_trips + fifty + no_work_recovery"
        if cfg.charge_min_trip_shortfall
        else "base + fifty + no_work_recovery (min-trip shortfall not charged)"
    )
    info.append([f"Customer_grand_baht ({cg_note})", customer_grand_baht])

    # --- Customer Summary ---
    cs = wb.create_sheet("Customer_Summary")
    cs.append(["Line", "รายการ", "บาท"])
    mx = sum_manual_extra_baht(cfg)
    cs.append(["A", "ค่าเที่ยวปกติ (GPS matched + เที่ยวเพิ่มจาก config)", base_baht])
    if mx:
        cs.append(["A2", "ในนั้น: เที่ยวเพิ่ม (manual_extra_trips ไม่มีใน GPS)", mx])
    mr = sum_manual_return_baht(cfg)
    if mr:
        cs.append(
            [
                "R",
                "ค่าขนส่งขากลับ (manual_return_trips — ไม่เพิ่มจำนวน matched)",
                mr,
            ]
        )
    if cfg.charge_min_trip_shortfall:
        b_line = f"เที่ยวขาดจาก commit {cfg.min_trips_per_truck} เที่ยว/คัน/วัน (min trips)"
    else:
        b_line = (
            f"ค่าชดเชยเที่ยวขาด (min {cfg.min_trips_per_truck}/คัน/วัน) — ไม่เก็บเงิน "
            f"(ใช้ชาร์จ {cfg.one_trip_surcharge_pct:.0f}% วันละ 1 เที่ยวแทน)"
        )
    cs.append(["B", b_line, min_trip_extra_baht])
    cs.append(["C", f"ชาร์จ {cfg.one_trip_surcharge_pct:.0f}% วันที่วิ่งได้ 1 เที่ยว (หลัง override)", fifty_total_baht])
    cs.append(
        [
            "D",
            "No-work recovery outbound 50pct (first matched trip that Dest_In day on recovery dates)",
            no_work_total_baht,
        ]
    )
    tot_lbl = (
        "Grand (A+B+C+D)"
        if cfg.charge_min_trip_shortfall
        else ("Grand (A+C+D+R)" if sum_manual_return_baht(cfg) else "Grand (A+C+D)")
    )
    cs.append(["TOTAL", tot_lbl, customer_grand_baht])

    # --- Customer: trips per day (matched, by Dest_In date) ---
    cpd = wb.create_sheet("Customer_Trips_Per_Day")
    cpd.append(["วันที่_Dest_In", "จำนวนเที่ยว_matched", "จำนวนรถ_มีเที่ยววันนั้น"])
    for r in cpd_rows:
        cpd.append([r["dest_date"], r["matched_trips"], r["active_trucks"]])

    # --- Audit Log (ชีตใหม่ — อธิบายเหตุผลการคิดเงินรายวัน/ทะเบียน) ---
    al = wb.create_sheet("Audit_Log")
    al.append([
        "Dest_In_date", "Plate", "Site",
        "เที่ยว", "เรท(฿)", "ค่าเที่ยว(฿)",
        f"+{cfg.one_trip_surcharge_pct:.0f}%(฿)", "ขากลับ(฿)", "รวมวันนี้(฿)",
        "เหตุผลการคิดเงิน",
    ])
    for r in audit_rows:
        al.append([
            r["dest_date"], r["plate"], r["site"],
            r["matched_trips"], r["trip_rate_baht"], r["base_line_baht"],
            r["fifty_pct_baht"], int(r.get("return_trip_baht", 0) or 0), r["customer_day_baht"],
            r["billing_note"],
        ])

    # --- Trip Detail ---
    td = wb.create_sheet("Trip_Detail")
    td.append([
        "Trip_Date", "Origin_Date", "Dest_Date", "Site", "Plate", "Device",
        "Origin_Row", "Dest_Row",
        "Origin_In", "Origin_Out", "Origin_Wait_h",
        "Dest_In", "Dest_Out",
        "Travel_h(OriginOut->DestIn)", "Dest_Wait_h", "Dest_Wait_customer_h", "Customer_idle_clip_h",
        "Total_Cycle_h", "Total_Cycle_customer_h",
        "Travel_Flag", "Billable_Trip", "Nw_outbound50_baht", "Return_manual_baht",
    ])
    firsts = first_matched_trip_by_plate_dest(trips)
    first_no_work = first_no_work_trip_by_plate_recovery_day(trips, cfg)
    ret_by_pd: dict[tuple[str, date], int] = {}
    for m in cfg.manual_return_trips:
        k = (str(m.plate), m.dest_date)
        ret_by_pd[k] = int(ret_by_pd.get(k, 0)) + int(m.amount_baht)
    for t in sorted(trips, key=lambda x: (x.dest_date, x.plate, x.d_in)):
        dw_c = customer_idle_clip_dest_wait_h(t, cfg)
        clip = max(0.0, t.dest_wait_h - dw_c)
        cyc_c = max(0.0, t.total_cycle_h - clip)
        td.append([
            t.trip_date, t.origin_date, t.dest_date,
            t.site, t.plate, t.device, t.o_row, t.d_row,
            t.o_in, t.o_out, round(t.origin_wait_h, 2),
            t.d_in, t.d_out,
            round(t.travel_h, 2), round(t.dest_wait_h, 2), round(dw_c, 2), round(clip, 2),
            round(t.total_cycle_h, 2), round(cyc_c, 2),
            t.travel_flag, 1, trip_no_work_outbound_baht(t, first_no_work, cfg),
            (
                int(ret_by_pd.get((str(t.plate), t.dest_date), 0))
                if firsts.get((t.plate, t.dest_date)) is not None
                and id(firsts.get((t.plate, t.dest_date))) == id(t)
                else 0
            ),
        ])

    # --- Unmatched Log ---
    um = wb.create_sheet("Unmatched_Log")
    um.append(
        ["Source", "Plate", "Device", "Row_No", "In", "Out", "Dwell_h", "Gap_to_next_In_h"]
    )
    for src, leg, _ in sorted(unmatched, key=lambda x: (x[2], x[1].t_in)):
        d_dw, g_gap = um_leg_dwell_gap_h(leg, leg_timeline_by_plate.get(leg.plate))
        um.append(
            [
                src,
                leg.plate,
                leg.device,
                leg.row_no,
                leg.t_in,
                leg.t_out,
                round(d_dw, 4),
                round(g_gap, 4) if g_gap is not None else "",
            ]
        )

    # --- Daily Activity ---
    da = wb.create_sheet("Daily_Activity")
    da.append([
        "Dest_In date", "Active trucks (all)", "Actual trips (all)",
        f"Commit min ({cfg.min_trips_per_truck}x trucks)", "Shortfall trips (all)",
        "BigC trucks", "BigC trips", "LCB trucks", "LCB trips",
    ])
    for d, s in daily_rows:
        da.append([
            d, s["trucks"], s["trips"], s["commit"], s["short"],
            len(s["bigc_p"]), s["bigc_t"], len(s["lcb_p"]), s["lcb_t"],
        ])

    # --- Daily Time 24h Check ---
    dt = wb.create_sheet("Daily_Time_24h_Check")
    dt.append([
        "Activity_Date", "Plate", "Site",
        "Matched_Cycle_h", "Matched_Origin_Wait_h", "Matched_Dest_Wait_h", "Matched_Travel_h",
        "Unmatched_Origin_h", "Unmatched_Dest_h",
        "Adjusted_Origin_Wait_h", "Adjusted_Dest_Wait_h",
        "Combined_h(AdjustedWait+Travel)", "Gap_vs_24_h",
    ])
    for d, plate, site, cycle_h, m_ow, m_dw, m_tr, um_ow, um_dw, ad_ow, ad_dw, comb, gap in daily_time:
        dt.append([
            d.isoformat(), plate, site,
            round(cycle_h, 2), round(m_ow, 2), round(m_dw, 2), round(m_tr, 2),
            round(um_ow, 2), round(um_dw, 2), round(ad_ow, 2), round(ad_dw, 2),
            round(comb, 2), round(gap, 2),
        ])

    # --- Plate DestDay ---
    pd_sheet = wb.create_sheet("Plate_DestDay")
    pd_sheet.append([
        "Dest_In_date", "Plate", "Site", "Matched_trips",
        "Trip_rate_baht", "Base_line_baht", "Fifty_pct_baht", "Return_trip_baht", "Customer_day_baht",
    ])
    for r in pday:
        pd_sheet.append([
            r["dest_date"], r["plate"], r["site"], r["matched_trips"],
            r["trip_rate_baht"], r["base_line_baht"], r["fifty_pct_baht"],
            int(r.get("return_trip_baht", 0) or 0),
            r["customer_day_baht"],
        ])

    # --- Surcharge 50% 1Trip ---
    lt = wb.create_sheet("Surcharge_50pct_1Trip")
    lt.append([
        "Dest_In_date", "Plate", "Site", "Fifty_kind",
        "Trips_that_day",
        "Auto_1trip_rule_Y/N", "Override_action", "Override_note",
        "Window_Origin_In", "Window_End",
        "Trip_rate_baht", f"Surcharge_baht_{cfg.one_trip_surcharge_pct:.0f}pct",
    ])
    for r in fifty_rows:
        lt.append([
            r["dest_date"], r["plate"], r["site"], str(r.get("fifty_kind", "")),
            r["trips_that_day"],
            "Y" if r["auto_1trip"] else "N",
            r.get("override_action", ""), r.get("override_note", ""),
            r.get("window_anchor", ""),
            r.get("window_end", ""),
            r["trip_rate_baht"], r["surcharge_baht"],
        ])

    mx = wb.create_sheet("Manual_Extra_Trips")
    mx.append(["Dest_In_date", "Plate", "Amount_baht", "Note"])
    for m in cfg.manual_extra_trips:
        mx.append([m.dest_date, m.plate, m.amount_baht, m.note])
    mr = wb.create_sheet("Manual_Return_Trips")
    mr.append(["Dest_In_date", "Plate", "Amount_baht", "Note"])
    for m in cfg.manual_return_trips:
        mr.append([m.dest_date, m.plate, m.amount_baht, m.note])
    nw = wb.create_sheet("NoWork_Outbound_50pct")
    nw.append(
        ["Dest_In_date", "Plate", "Site", "Dest_Row", "Trip_rate_baht", "Surcharge_baht_50pct", "Note"]
    )
    for r in no_work_rows:
        nw.append(
            [
                r["dest_date"],
                r["plate"],
                r["site"],
                r["d_row"],
                r["trip_rate_baht"],
                r["surcharge_baht"],
                r.get("note", ""),
            ]
        )
    ph = wb.create_sheet("Phantom_Trip_Candidates")
    ph.append(
        ["Plate", "Calendar_date", "Origin_hours", "Suggest_full_trip_baht", "Note"]
    )
    for r in phantom_rows:
        ph.append(
            [
                r["plate"],
                r["calendar_date"],
                r["origin_hours_on_day"],
                r["suggest_full_trip_baht"],
                r.get("note", ""),
            ]
        )
    hi = wb.create_sheet("Hints_DoubleOrigin")
    hi.append(["Plate", "Calendar_date", "UM_Origin_segments", "Note"])
    for r in hint_rows:
        hi.append(
            [r["plate"], r["calendar_date"], r["um_origin_segments"], r.get("note", "")]
        )

    beautify_oatside_workbook(wb)
    wb.save(path)


# ---------------------------------------------------------------------------
# HTML helpers
# ---------------------------------------------------------------------------

def esc(x) -> str:
    return html_module.escape(str(x), quote=True)


_TRIPS_FILTER_JS = (
    "<script>(function(){"
    "var sel=document.getElementById('tripsPlateFilter');"
    "var qel=document.getElementById('tripsPlateQuery');"
    "var tb=document.querySelector('#tripsAllTable tbody');"
    "if(!tb)return;"
    "function run(){"
    "var v=sel?(sel.value||'').trim():'';"
    "var q=qel?(qel.value||'').trim().toLowerCase():'';"
    "var rows=tb.querySelectorAll('tr');"
    "for(var i=0;i<rows.length;i++){"
    "var r=rows[i];"
    "var p=(r.getAttribute('data-plate')||'');"
    "var pok=!v||p===v;"
    "var qok=!q||p.toLowerCase().indexOf(q)>=0;"
    "r.style.display=(pok&&qok)?'':'none';"
    "}"
    "}"
    "if(sel)sel.addEventListener('change',run);"
    "if(qel)qel.addEventListener('input',run);"
    "})();</script>"
)


def html_fifty_surcharge_badge(fr: dict, cfg: OatsideConfig) -> str:
    """Badge: ตีเปล่า (เฉพาะที่ mark) vs ค่าเสียเวลา (+50%% / +100%% รวมข้ามคืน)."""
    amt = int(fr.get("surcharge_baht", 0) or 0)
    if amt <= 0:
        return ""
    rate = int(fr.get("trip_rate_baht", 0) or 0)
    kind = str(fr.get("fifty_kind") or "")
    pct = float(cfg.one_trip_surcharge_pct)
    if kind == "blank_run":
        label = f"ตีเปล่า +{pct:.0f}%"
        cls = "blankrun"
    elif kind == "no_work_outbound":
        label = f"ตีเปล่า +{pct:.0f}%"
        cls = "blankrun"
    elif kind == "midnight_full" or (not kind and rate > 0 and amt >= rate):
        label = "ค่าเสียเวลา +100%"
        cls = "fulltrip"
    elif kind == "midnight_pct":
        label = f"ค่าเสียเวลา +{pct:.0f}%"
        cls = "dwell"
    elif kind in ("origin24h", "downtime_dest", "downtime_origin_day"):
        label = f"ค่าเสียเวลา +{pct:.0f}%"
        cls = "dwell"
    else:
        if rate > 0 and amt >= rate:
            label = "ค่าเสียเวลา +100%"
            cls = "fulltrip"
        else:
            label = f"ค่าเสียเวลา +{pct:.0f}%"
            cls = "dwell"
    return f"<span class='badge {cls}'>{label} ฿{amt:,}</span>"


def fmt_h(x: float) -> str:
    return f"{x:.2f}".rstrip("0").rstrip(".")


def fmt_hm(x: float) -> str:
    sign = "-" if x < 0 else ""
    total_minutes = int(round(abs(x) * 60))
    hh = total_minutes // 60
    mm = total_minutes % 60
    return f"{sign}{hh}.{mm:02d}"


def unmatched_merged_trip_one_row_html(
    src: str,
    leg: Leg,
    *,
    dwell_h: float,
    gap_h: float | None,
    include_plate_link: bool = True,
    include_plate_column: bool = True,
) -> str:
    """One <tr> aligned with trip_row (full) or trip_row_plate (no plate column)."""
    dash = "—"
    tag = "UM-O" if src == "Origin" else "UM-D"
    badge = f"<span class='badge abn'>{tag}</span> "
    site = site_for_plate(leg.plate)
    site_html = f"<span class='badge {'bigc' if site == 'BigC' else 'lcb'}'>{site}</span>"
    if include_plate_column:
        if include_plate_link:
            plate_html = f"{badge}<a href='plates/{esc(leg.plate)}.html'>{esc(leg.plate)}</a>"
        else:
            plate_html = f"{badge}{esc(leg.plate)}"
        site_plate = f"<td>{site_html}</td><td>{plate_html}</td>"
    else:
        site_plate = f"<td>{site_html} {badge}</td>"
    if src == "Origin":
        od, dd = leg.t_in.date(), dash
        oi, oo = leg.t_in, leg.t_out
        di, do = dash, dash
    else:
        od, dd = dash, leg.t_in.date()
        oi, oo = dash, dash
        di, do = leg.t_in, leg.t_out
    return (
        f"<tr class='um' data-plate='{esc(leg.plate)}'><td>{od}</td><td>{dd}</td>{site_plate}"
        f"<td>{oi}</td><td>{oo}</td><td>{di}</td><td>{do}</td>"
        f"<td>{dash}</td><td>{dash}</td><td>{dash}</td>"
        f"<td>{fmt_hm(dwell_h)}</td><td>{fmt_hm(gap_h) if gap_h is not None else dash}</td>"
        f"<td>{dash}</td><td>{dash}</td><td>{dash}</td><td>{dash}</td><td>{dash}</td></tr>"
    )



def sum_manual_extra_baht(cfg: OatsideConfig) -> int:
    return sum(m.amount_baht for m in cfg.manual_extra_trips)


def merge_manual_extra_into_pday(pday_rows: list[dict], cfg: OatsideConfig) -> None:
    for m in cfg.manual_extra_trips:
        found = False
        for r in pday_rows:
            if str(r["plate"]) == m.plate and r["dest_date"] == m.dest_date:
                r["base_line_baht"] = int(r["base_line_baht"]) + m.amount_baht
                r["customer_day_baht"] = int(r["customer_day_baht"]) + m.amount_baht
                r["matched_trips"] = int(r["matched_trips"]) + 1
                tag = esc(m.note) if m.note else "เที่ยวเพิ่ม (ไม่มีใน GPS)"
                badge = (
                    f"<span class='badge manual-extra' title='{tag}'>"
                    f"เที่ยวเพิ่ม +{m.amount_baht:,}฿</span>"
                )
                prev = (r.get("fifty_badge_html") or "").strip()
                r["fifty_badge_html"] = (prev + " " + badge).strip() if prev else badge
                found = True
                break
        if not found:
            rate = trip_rate_baht(m.dest_date, cfg)
            tag = esc(m.note) if m.note else "เที่ยวเพิ่ม (ไม่มีใน GPS)"
            badge = (
                f"<span class='badge manual-extra' title='{tag}'>"
                f"เที่ยวเพิ่ม +{m.amount_baht:,}฿</span>"
            )
            pday_rows.append(
                {
                    "dest_date": m.dest_date,
                    "plate": m.plate,
                    "site": site_for_plate(m.plate),
                    "matched_trips": 1,
                    "trip_rate_baht": rate,
                    "base_line_baht": m.amount_baht,
                    "fifty_pct_baht": 0,
                    "fifty_badge_html": badge,
                    "customer_day_baht": m.amount_baht,
                }
            )
    pday_rows.sort(key=lambda r: (r["dest_date"], str(r["plate"])))


def merge_manual_extra_into_audit(audit_rows: list[dict], cfg: OatsideConfig) -> None:
    for m in cfg.manual_extra_trips:
        hit = False
        for r in audit_rows:
            if str(r["plate"]) != m.plate or r.get("dest_date") != m.dest_date:
                continue
            r["base_line_baht"] = int(r["base_line_baht"]) + m.amount_baht
            r["customer_day_baht"] = int(r["customer_day_baht"]) + m.amount_baht
            r["matched_trips"] = int(r["matched_trips"]) + 1
            extra = (
                f" | เที่ยวเพิ่ม (ไม่มีใน GPS): {m.note} (+{m.amount_baht:,}฿)"
                if m.note
                else f" | เที่ยวเพิ่ม (ไม่มีใน GPS) +{m.amount_baht:,}฿"
            )
            r["billing_note"] = str(r.get("billing_note", "")) + extra
            hit = True
            break
        if hit:
            continue
        rate = trip_rate_baht(m.dest_date, cfg)
        note = (
            f"เที่ยวเพิ่ม (ไม่มีใน GPS): {m.note} (+{m.amount_baht:,}฿)"
            if m.note
            else f"เที่ยวเพิ่ม (ไม่มีใน GPS) +{m.amount_baht:,}฿"
        )
        audit_rows.append(
            {
                "origin_day": m.dest_date,
                "dest_date": m.dest_date,
                "plate": m.plate,
                "site": site_for_plate(m.plate),
                "matched_trips": 1,
                "trip_rate_baht": rate,
                "base_line_baht": m.amount_baht,
                "fifty_pct_baht": 0,
                "customer_day_baht": m.amount_baht,
                "billing_note": note,
            }
        )
    audit_rows.sort(key=lambda r: (r.get("origin_day", r["dest_date"]), str(r["plate"])))


def apply_manual_extra_to_cpd(cpd_rows: list[dict], cfg: OatsideConfig) -> None:
    by_d = {r["dest_date"]: r for r in cpd_rows}
    for m in cfg.manual_extra_trips:
        if m.dest_date in by_d:
            by_d[m.dest_date]["matched_trips"] = int(by_d[m.dest_date]["matched_trips"]) + 1
        else:
            cpd_rows.append(
                {"dest_date": m.dest_date, "matched_trips": 1, "active_trucks": 1}
            )
            by_d[m.dest_date] = cpd_rows[-1]
    cpd_rows.sort(key=lambda r: r["dest_date"])




def sum_manual_return_baht(cfg: OatsideConfig) -> int:
    return sum(m.amount_baht for m in cfg.manual_return_trips)


def merge_manual_return_into_pday(pday_rows: list[dict], cfg: OatsideConfig) -> None:
    for m in cfg.manual_return_trips:
        found = False
        for r in pday_rows:
            if str(r["plate"]) == m.plate and r["dest_date"] == m.dest_date:
                prev = int(r.get("return_trip_baht", 0) or 0)
                r["return_trip_baht"] = prev + int(m.amount_baht)
                r["customer_day_baht"] = int(r["customer_day_baht"]) + int(m.amount_baht)
                tag = esc(m.note) if m.note else "ค่าขนส่งขากลับ (manual)"
                badge = (
                    f"<span class='badge return-trip' title='{tag}'>"
                    f"ขากลับ +{m.amount_baht:,}฿</span>"
                )
                prev_b = (r.get("fifty_badge_html") or "").strip()
                r["fifty_badge_html"] = (prev_b + " " + badge).strip() if prev_b else badge
                found = True
                break
        if not found:
            rate = trip_rate_baht(m.dest_date, cfg)
            tag = esc(m.note) if m.note else "ค่าขนส่งขากลับ (manual)"
            badge = (
                f"<span class='badge return-trip' title='{tag}'>"
                f"ขากลับ +{m.amount_baht:,}฿</span>"
            )
            pday_rows.append(
                {
                    "dest_date": m.dest_date,
                    "plate": m.plate,
                    "site": site_for_plate(m.plate),
                    "matched_trips": 0,
                    "trip_rate_baht": rate,
                    "base_line_baht": 0,
                    "fifty_pct_baht": 0,
                    "fifty_badge_html": badge,
                    "return_trip_baht": int(m.amount_baht),
                    "customer_day_baht": int(m.amount_baht),
                }
            )
    pday_rows.sort(key=lambda r: (r["dest_date"], str(r["plate"])))


def merge_manual_return_into_audit(audit_rows: list[dict], cfg: OatsideConfig) -> None:
    for m in cfg.manual_return_trips:
        hit = False
        for r in audit_rows:
            if str(r["plate"]) != m.plate or r.get("dest_date") != m.dest_date:
                continue
            prev = int(r.get("return_trip_baht", 0) or 0)
            r["return_trip_baht"] = prev + int(m.amount_baht)
            r["customer_day_baht"] = int(r["customer_day_baht"]) + int(m.amount_baht)
            extra = (
                f" | ขากลับ (manual): {m.note} (+{m.amount_baht:,}฿)"
                if m.note
                else f" | ขากลับ (manual) +{m.amount_baht:,}฿"
            )
            r["billing_note"] = str(r.get("billing_note", "")) + extra
            hit = True
            break
        if hit:
            continue
        rate = trip_rate_baht(m.dest_date, cfg)
        note = (
            f"ขากลับ (manual): {m.note} (+{m.amount_baht:,}฿)"
            if m.note
            else f"ขากลับ (manual) +{m.amount_baht:,}฿"
        )
        audit_rows.append(
            {
                "origin_day": m.dest_date,
                "dest_date": m.dest_date,
                "plate": m.plate,
                "site": site_for_plate(m.plate),
                "matched_trips": 0,
                "trip_rate_baht": rate,
                "base_line_baht": 0,
                "fifty_pct_baht": 0,
                "return_trip_baht": int(m.amount_baht),
                "customer_day_baht": int(m.amount_baht),
                "billing_note": note,
            }
        )
    audit_rows.sort(key=lambda r: (r.get("origin_day", r["dest_date"]), str(r["plate"])))

def _tr_prepend_day_band(html: str, day: date) -> str:
    """Zebra by calendar day (Origin_In day for matched; UM-O uses leg time; UM-D has no Origin on row — uses leg time) — subtle band in CSS."""
    band = f"day-band-{day.toordinal() % 2}"
    if html.startswith("<tr class='"):
        return html.replace("<tr class='", f"<tr class='{band} ", 1)
    if html.startswith("<tr>"):
        return html.replace("<tr>", f"<tr class='{band}'>", 1)
    return html


def interleaved_matched_unmatched_rows_html(
    trips: list[Trip],
    unmatched: list[tuple[str, Leg, str]],
    trip_row_cb: Callable[[Trip], str],
    *,
    plate: str | None = None,
    include_plate_link: bool = True,
    include_plate_column: bool = True,
    leg_timeline_by_plate: dict[str, list[Leg]] | None = None,
) -> str:
    """Sort matched by Origin_In time; unmatched by leg t_in (UM-O=Origin, UM-D=Dest)."""
    rows: list[tuple[datetime, tuple[Any, ...], str]] = []
    for t in trips:
        if plate is not None and t.plate != plate:
            continue
        day = t.o_in.date()
        html = _tr_prepend_day_band(trip_row_cb(t), day)
        rows.append((t.o_in, (0, t.plate, t.d_row or "", t.o_row or ""), html))
    for src, leg, _mp in unmatched:
        if plate is not None and leg.plate != plate:
            continue
        _dw, _gp = um_leg_dwell_gap_h(
            leg, leg_timeline_by_plate.get(leg.plate) if leg_timeline_by_plate else None
        )
        um_html = unmatched_merged_trip_one_row_html(
            src,
            leg,
            dwell_h=_dw,
            gap_h=_gp,
            include_plate_link=include_plate_link,
            include_plate_column=include_plate_column,
        )
        um_html = _tr_prepend_day_band(um_html, leg.t_in.date())
        kind = 1 if src == "Origin" else 2
        rows.append((leg.t_in, (kind, leg.plate, src, leg.row_no), um_html))
    rows.sort(key=lambda x: (x[0], x[1]))
    return "".join(r[2] for r in rows)


# ---------------------------------------------------------------------------
# HTML export
# ---------------------------------------------------------------------------

def write_html(
    report_dir: Path,
    origin_label: str,
    trips: list[Trip],
    daily_rows: list[tuple[date, dict]],
    daily_time: list[tuple],
    actual: int,
    commit: int,
    short: int,
    extra: int,
    bc: tuple,
    fifty_rows: list[dict],
    fifty_total_baht: int,
    grand_extra_baht: int,
    base_baht: int,
    customer_grand_baht: int,
    pday_rows: list[dict],
    audit_rows: list[dict],
    unmatched: list[tuple[str, Leg, str]],
    nw_total_baht: int,
    cfg: OatsideConfig,
    cpd_rows: list[dict],
    leg_timeline_by_plate: dict[str, list[Leg]],
) -> None:
    bc_a, bc_c, bc_s, bc_e, lc_a, lc_c, lc_s, lc_e = bc
    thr = iqr_threshold([t.travel_h for t in trips])
    abn = [t for t in trips if t.travel_flag]
    plates = sorted({t.plate for t in trips})
    fifty_by_lists: dict[tuple[str, date], list[dict]] = defaultdict(list)
    for r in fifty_rows:
        fifty_by_lists[(r["plate"], r["dest_date"])].append(r)
    fifty_origin_lists: dict[tuple[str, date], list[dict]] = defaultdict(list)
    for r in fifty_rows:
        if "origin_day" in r:
            fifty_origin_lists[(r["plate"], r["origin_day"])].append(r)
    firsts = first_matched_trip_by_plate_dest(trips)
    first_no_work = first_no_work_trip_by_plate_recovery_day(trips, cfg)
    ret_by_pd: dict[tuple[str, date], int] = {}
    for m in cfg.manual_return_trips:
        k = (str(m.plate), m.dest_date)
        ret_by_pd[k] = int(ret_by_pd.get(k, 0)) + int(m.amount_baht)
    _um_rows: list[str] = []
    for src, leg, _ in sorted(unmatched, key=lambda x: x[1].t_in):
        _dwell, _gap = um_leg_dwell_gap_h(leg, leg_timeline_by_plate.get(leg.plate))
        _gap_cell = fmt_hm(_gap) if _gap is not None else "—"
        _um_rows.append(
            f"<tr><td><span class='badge abn'>{'UM-O' if src == 'Origin' else 'UM-D'}</span></td>"
            f"<td><a href='plates/{esc(leg.plate)}.html'>{esc(leg.plate)}</a></td>"
            f"<td>{leg.t_in}</td><td>{leg.t_out}</td>"
            f"<td class='note'>{fmt_hm(_dwell)}</td><td class='note'>{_gap_cell}</td>"
            f"<td class='note'>{'Origin ไม่มีคู่' if src == 'Origin' else 'Dest ไม่มีคู่'}</td></tr>"
        )
    um_section_html = "".join(_um_rows) or "<tr><td colspan=7 class='note'>ไม่มี Unmatched</td></tr>"
    sub = (
        f"สร้าง {datetime.now():%Y-%m-%d %H:%M} | ต้นทาง: {esc(Path(origin_label).name)} | "
        f"เรท: {config_rate_summary(cfg)} ฿/เที่ยว | "
        f"min {cfg.min_trips_per_truck} เที่ยว/คัน/วัน | "
        f"+{cfg.one_trip_surcharge_pct:.0f}% วันที่วิ่ง 1 เที่ยว | "
        f"max travel {cfg.max_travel_h}h"
    )
    if not cfg.charge_min_trip_shortfall:
        sub += " | ไม่เก็บเงินค่าชดเชยเที่ยวขาด (min trips) — ใช้ชาร์จ % วันละ 1 เที่ยวแทน"


    _hi_o = float(getattr(cfg, "highlight_origin_wait_h", 8.0))
    _hi_d = float(getattr(cfg, "highlight_dest_wait_h", 8.0))

    def _td_wait_h(val: float, th: float, dest: bool) -> str:
        cls = "wait-hi-dest" if dest else "wait-hi"
        if val >= th:
            lab = "ปลายทาง" if dest else "ต้นทาง"
            return f"<td class='{cls}' title='รอ{lab} ≥ {th:g} ชม. (ตรวจพิจารณา)'>{fmt_hm(val)}</td>"
        return f"<td>{fmt_hm(val)}</td>"

    def trip_row(t: Trip) -> str:
        ab = " <span class='badge abn'>ABNORMAL</span>" if t.travel_flag else ""
        ft0 = firsts.get((t.plate, t.dest_date))
        ret_amt = (
            int(ret_by_pd.get((str(t.plate), t.dest_date), 0))
            if ft0 is not None and id(ft0) == id(t)
            else 0
        )
        money = trip_row_pricing_cells(
            t,
            firsts=firsts,
            first_no_work=first_no_work,
            fifty_by_lists=fifty_by_lists,
            cfg=cfg,
            return_baht=ret_amt,
        )
        return (
            f"<tr data-plate='{esc(t.plate)}'><td>{t.origin_date}</td><td>{t.dest_date}</td>"
            f"<td><span class='badge {'bigc' if t.site=='BigC' else 'lcb'}'>{t.site}</span></td>"
            f"<td><a href='plates/{esc(t.plate)}.html'>{esc(t.plate)}</a>{ab}</td>"
            f"<td>{t.o_in}</td><td>{t.o_out}</td><td>{t.d_in}</td><td>{t.d_out}</td>"
            f"{_td_wait_h(t.origin_wait_h, _hi_o, False)}<td>{fmt_hm(t.travel_h)}</td>{_td_wait_h(t.dest_wait_h, _hi_d, True)}"
            f"<td>—</td><td>—</td>"
            f"{money}</tr>"
        )

    merged_all_rows = interleaved_matched_unmatched_rows_html(
        trips,
        unmatched,
        trip_row,
        plate=None,
        include_plate_link=True,
        include_plate_column=True,
        leg_timeline_by_plate=leg_timeline_by_plate,
    )

    def trip_row_plate(t: Trip) -> str:
        ab = " <span class='badge abn'>ABNORMAL</span>" if t.travel_flag else ""
        ft0 = firsts.get((t.plate, t.dest_date))
        ret_amt = (
            int(ret_by_pd.get((str(t.plate), t.dest_date), 0))
            if ft0 is not None and id(ft0) == id(t)
            else 0
        )
        money = trip_row_pricing_cells(
            t,
            firsts=firsts,
            first_no_work=first_no_work,
            fifty_by_lists=fifty_by_lists,
            cfg=cfg,
            return_baht=ret_amt,
        )
        return (
            f"<tr data-plate='{esc(t.plate)}'><td>{t.origin_date}</td><td>{t.dest_date}</td><td>{t.site}{ab}</td>"
            f"<td>{t.o_in}</td><td>{t.o_out}</td><td>{t.d_in}</td><td>{t.d_out}</td>"
            f"{_td_wait_h(t.origin_wait_h, _hi_o, False)}<td>{fmt_hm(t.travel_h)}</td>{_td_wait_h(t.dest_wait_h, _hi_d, True)}"
            f"<td>—</td><td>—</td>"
            f"{money}</tr>"
        )

    daily_act_rows_html = "".join(
        f"<tr><td>{d}</td><td>{s['trucks']}</td><td>{s['trips']}</td><td>{s['commit']}</td><td>{s['short']}</td>"
        f"<td>{len(s['bigc_p'])}</td><td>{s['bigc_t']}</td><td>{len(s['lcb_p'])}</td><td>{s['lcb_t']}</td></tr>"
        for d, s in daily_rows
    )

    tpd_rows_html = "".join(
        f"<tr><td>{r['dest_date']}</td><td>{r['matched_trips']}</td><td>{r['active_trucks']}</td></tr>"
        for r in cpd_rows
    ) or "<tr><td colspan=3>ไม่มีข้อมูล</td></tr>"

    dt_rows_html = "".join(
        f"<tr><td>{d}</td><td><a href='plates/{esc(p)}.html'>{esc(p)}</a></td>"
        f"<td><span class='badge {'bigc' if site=='BigC' else 'lcb'}'>{site}</span></td>"
        f"<td>{fmt_hm(cycle_h)}</td><td>{fmt_hm(m_ow)}</td><td>{fmt_hm(m_dw)}</td><td>{fmt_hm(m_tr)}</td>"
        f"<td>{fmt_hm(um_ow)}</td><td>{fmt_hm(um_dw)}</td><td>{fmt_hm(ad_ow)}</td><td>{fmt_hm(ad_dw)}</td><td>{fmt_hm(comb)}</td></tr>"
        for d, p, site, cycle_h, m_ow, m_dw, m_tr, um_ow, um_dw, ad_ow, ad_dw, comb, gap in daily_time
    )

    abn_rows_html = "".join(
        f"<tr><td>{t.origin_date}</td><td>{t.dest_date}</td><td><a href='plates/{esc(t.plate)}.html'>{esc(t.plate)}</a></td>"
        f"<td>{t.site}</td><td>{t.o_out}</td><td>{t.d_in}</td><td>{fmt_hm(t.travel_h)}</td></tr>"
        for t in abn
    )

    lt_rows_html = "".join(
        f"<tr><td>{r['dest_date']}</td><td><a href='plates/{esc(r['plate'])}.html'>{esc(r['plate'])}</a></td>"
        f"<td><span class='badge {'bigc' if r['site']=='BigC' else 'lcb'}'>{r['site']}</span></td>"
        f"<td class='note'>{esc(str(r.get('fifty_kind','')))}</td>"
        f"<td>{r['trips_that_day']}</td><td>{'Y' if r['auto_1trip'] else 'N'}</td>"
        f"<td>{esc(r.get('override_action',''))}</td><td>{esc(r.get('override_note',''))}</td>"
        f"<td>{esc(r.get('window_anchor',''))}</td><td>{esc(r.get('window_end',''))}</td>"
        f"<td>{r['trip_rate_baht']:,}</td><td class='money'>{r['surcharge_baht']:,}</td>"
        f"<td>{html_fifty_surcharge_badge(r, cfg)}</td></tr>"
        for r in fifty_rows
    )

    # Audit table — สรุปเหตุผลรายวัน/ทะเบียน (origin_day เมื่อใช้ mode นั้น)
    audit_html = "".join(
        f"<tr><td>{r.get('origin_day', r['dest_date'])}</td>"
        f"<td><a href='plates/{esc(r['plate'])}.html'>{esc(r['plate'])}</a></td>"
        f"<td><span class='badge {'bigc' if r['site']=='BigC' else 'lcb'}'>{r['site']}</span></td>"
        f"<td>{r['matched_trips']}</td><td>{r['trip_rate_baht']:,}</td>"
        f"<td>{r['base_line_baht']:,}</td>"
        f"<td class='{'money' if r['fifty_pct_baht'] else ''}'>{r['fifty_pct_baht']:,}</td>"
        f"<td class='money'>{int(r.get('return_trip_baht',0) or 0):,}</td>"
        f"<td class='money'>{r['customer_day_baht']:,}</td>"
        f"<td class='note'>{esc(r['billing_note'])}</td></tr>"
        for r in audit_rows
    )

    css = (
        "body{font-family:Segoe UI,Tahoma,sans-serif;margin:24px;background:#f4f7fb;color:#152235}"
        "a{color:#0b57d0;text-decoration:none}a:hover{text-decoration:underline}"
        ".h1{font-size:30px;font-weight:700;margin-bottom:4px}.sub{color:#4b5b74;margin-bottom:16px;font-size:13px}"
        ".grid{display:grid;grid-template-columns:repeat(4,minmax(180px,1fr));gap:10px;margin-bottom:14px}"
        ".card{background:#fff;border-radius:10px;padding:12px;box-shadow:0 2px 8px rgba(16,24,40,.08)}"
        ".label{font-size:12px;color:#63758f}.value{font-size:28px;font-weight:700}"
        ".money{color:#0d6b3c}.warn{color:#b54708}"
        ".panel{background:#fff;border-radius:10px;padding:14px;box-shadow:0 2px 8px rgba(16,24,40,.08);margin-bottom:14px}"
        "table{width:100%;border-collapse:collapse;font-size:14px}"".table-scroll{overflow:auto;max-height:72vh;border:1px solid #e6ebf2;border-radius:8px;margin-top:8px}"".table-scroll thead th{position:sticky;top:0;z-index:4;background:#eef3fa;box-shadow:0 1px 0 #c5d0e0}"
        "th,td{padding:8px;border-bottom:1px solid #e6ebf2;text-align:left}th{background:#eef3fa}"
        ".badge{display:inline-block;padding:2px 8px;border-radius:999px;font-size:12px;font-weight:700;margin:0 6px 4px 0}"
        ".bigc{background:#dfebff;color:#0a4da1}.lcb{background:#e3f5e9;color:#0f6a3b}"
        ".abn{background:#ffe9e9;color:#b42318}.nav{margin-bottom:12px}"
        ".note{color:#4b5b74;font-size:13px}.wait-hi{background:#fff3cd;font-weight:600}.wait-hi-dest{background:#ffe0b2;font-weight:600}"
        ".fulltrip{background:#e3f2fd;color:#0d47a1}.blankrun{background:#ede7f6;color:#4a148c}.dwell{background:#fff3e0;color:#bf360c}"
        "tr.um td{color:#5a3b00}"
        ".manual-extra{background:#ede7f6;color:#4a148c;font-weight:600}.return-trip{background:#e8f5e9;color:#1b5e20;font-weight:600}"
        "tr.day-band-0 td{background:#fafcfe}tr.day-band-1 td{background:#e9f1fa}tr.day-band-0 td.wait-hi{background:#fff1cc;font-weight:600}tr.day-band-1 td.wait-hi{background:#ffecc4;font-weight:600}tr.day-band-0 td.wait-hi-dest{background:#ffe8c8;font-weight:600}tr.day-band-1 td.wait-hi-dest{background:#ffdfba;font-weight:600}""details.section-fold{margin-bottom:10px}""summary.section-sum{cursor:pointer;padding:10px 14px;background:#fff;border-radius:10px;font-weight:600;margin-bottom:6px;display:block;box-shadow:0 2px 8px rgba(16,24,40,.08);list-style:none}""summary.section-sum::-webkit-details-marker{display:none}"".filter-bar{display:flex;gap:10px;align-items:center;flex-wrap:wrap;margin:8px 0 14px}"".filter-bar label{font-size:13px;color:#4b5b74}"".filter-bar select,.filter-bar input[type=search]{font:inherit;padding:6px 10px;border-radius:8px;border:1px solid #c5d0e0;background:#fff;min-width:160px}""summary.section-sum-row{display:flex!important;width:100%;box-sizing:border-box;justify-content:space-between;align-items:center;gap:12px;list-style:none}""summary.section-sum-row .sum-main{flex:1 1 auto;min-width:0;text-align:left}""summary.section-sum-row .sum-dl{margin-left:auto;flex:0 0 auto}"".xlsx-dl{font-size:12px;font-weight:700;color:#0b57d0;padding:5px 10px;border-radius:8px;border:1px solid #b8cff4;background:#eef5ff;white-space:nowrap}"".hero-trips{display:flex;flex-wrap:wrap;align-items:center;justify-content:space-between;gap:14px;background:linear-gradient(135deg,#e8f1ff,#ffffff);border:1px solid #c5d0e0;border-radius:12px;padding:16px 18px;margin:12px 0 16px}"".hero-copy{max-width:720px}"".hero-tag{display:inline-block;font-size:11px;font-weight:700;color:#0b57d0;background:#e3eeff;border-radius:999px;padding:2px 10px;margin-bottom:6px}"".hero-title{font-size:20px;font-weight:800;color:#12243b;margin-bottom:4px}"".hero-sub{color:#4b5b74;font-size:13px;line-height:1.45}"".btn-primary{display:inline-block;padding:12px 18px;border-radius:10px;background:#0b57d0;color:#fff;font-weight:800;box-shadow:0 4px 12px rgba(11,87,208,.22)}"".btn-primary:hover{filter:brightness(1.05)}"".nav-secondary{margin:0 0 12px;font-size:13px;color:#4b5b74}"".panel-title-row{display:flex;justify-content:space-between;align-items:flex-start;gap:12px;flex-wrap:wrap}"".panel-title-row h3{margin:0}"".h1 .trips-tag{font-size:13px;font-weight:800;color:#0b57d0;margin-left:8px;vertical-align:middle}"".trips-lead{color:#4b5b74;font-size:14px;margin:-2px 0 10px}"
    )

    def _xlsx_dl(fname: str, short: str) -> str:
        return (
            "<a class='xlsx-dl' href='exports/"
            + str(fname)
            + "' download onclick='event.stopPropagation()'>ดาวน์โหลด "
            + html_module.escape(str(short), quote=False)
            + "</a>"
        )

    idx = f"""<!doctype html><html><head><meta charset='utf-8'><meta name='viewport' content='width=device-width,initial-scale=1'>
<title>Oatside report</title><style>{css}</style></head><body>
<div class='h1'>Oatside → P&amp;G สรุปรายงาน</div>
<div class='sub'>{sub}</div>
<div class='hero-trips'><div class='hero-copy'><div class='hero-tag'>แนะนำสำหรับลูกค้า</div><div class='hero-title'>เริ่มจากรายการเที่ยวทั้งหมด</div><div class='hero-sub'>เวลาเข้า-ออกครบ · ค่าขนส่ง / ส่วนเพิ่ม / ขากลับ — กรองทะเบียนได้ · ดาวน์โหลด Excel รายเที่ยวละเอียดได้จากปุ่มขวาบนหัวตารางในหน้าเที่ยวทั้งหมด</div></div><a class='btn-primary' href='trips.html'>เปิดเที่ยวทั้งหมด</a></div><div class='nav-secondary'><a href='trips.html'>ดูเที่ยวทั้งหมด</a> · <a href='../../../Oatside/Oatside_PG_Trip_Summary_By_Site.xlsx'>ดาวน์โหลด Excel รวมทุกชีต</a></div>
<div class='grid'>
<div class='card'><div class='label'>ค่าเที่ยวปกติ (A)</div><div class='value money'>{base_baht:,}</div></div>
<div class='card'><div class='label'>ชาร์จเสริม ตีเปล่า/เสียเวลา/ข้ามคืน (C)</div><div class='value money'>{fifty_total_baht:,}</div></div>
<div class='card'><div class='label'>No-work Recovery +50% (D)</div><div class='value money'>{nw_total_baht:,}</div></div>
<div class='card'><div class='label'>รวมลูกค้า</div><div class='value money'>{customer_grand_baht:,}</div></div>
</div>
<details class='section-fold'><summary class='section-sum section-sum-row'><span class='sum-main'>(1) จำนวนเที่ยวต่อวัน (matched Dest_In)</span><span class='sum-dl'>{_xlsx_dl('01_CPD_MatchedTripsPerDay.xlsx', 'ตาราง (1)')}</span></summary>
<div class='panel'>
<p class='sub'>นับตาม Dest_In · รวมทุกทะเบียน · เฉพาะเที่ยวที่จับคู่แล้ว</p>
<table><thead><tr><th>วันที่</th><th>จำนวนเที่ยว</th><th>จำนวนรถ</th></tr></thead><tbody>
{tpd_rows_html}
</tbody></table></div>
</details>
<details class='section-fold'><summary class='section-sum section-sum-row'><span class='sum-main'>(2) เดลี่รถทุกคัน — Dest_In × ทะเบียน</span><span class='sum-dl'>{_xlsx_dl('02_Plate_DestDay_Daily.xlsx', 'ตาราง (2)')}</span></summary>
<div class='panel'>
<p class='sub'>เรท: {config_rate_summary(cfg)} ฿/เที่ยว · คอลัมน์ส่วนเพิ่มแสดงได้หลายป้ายในวันเดียวกัน (เว้นวรรค) — ตีเปล่า = No-work recovery หรือ mark override; ค่าเสียเวลา = fifty; ข้ามคืนเต็มเที่ยว = +100% (หลัง override) · Policy: recovery-day บวกคู่กับ fifty หากมี (2026-05-01){'<br>ตาราง (2) นับตาม Dest_In · <b>Audit Log ด้านล่างคิดตาม วันงาน (Origin_In)</b>' if cfg.use_origin_day_fifty else ''}</p>
<table><thead><tr><th>วันที่</th><th>ทะเบียน</th><th>Site</th><th>เที่ยว</th><th>เรท(฿)</th><th>ค่าเที่ยว(฿)</th><th>ส่วนเพิ่ม (฿)</th><th>ขากลับ(฿)</th><th>รวมวัน(฿)</th></tr></thead><tbody>
{"".join(f"<tr><td>{r['dest_date']}</td><td><a href='plates/{esc(r['plate'])}.html'>{esc(r['plate'])}</a></td><td><span class='badge {'bigc' if r['site']=='BigC' else 'lcb'}'>{r['site']}</span></td><td>{r['matched_trips']}</td><td>{r['trip_rate_baht']:,}</td><td>{r['base_line_baht']:,}</td><td>{(r['fifty_badge_html'] if r.get('fifty_badge_html') else f"<span class='money'>{r['fifty_pct_baht']:,}</span>")}</td><td class='money'>{int(r.get('return_trip_baht',0) or 0):,}</td><td class='money'>{r['customer_day_baht']:,}</td></tr>" for r in pday_rows) or "<tr><td colspan=9>ไม่มีข้อมูล</td></tr>"}
</tbody></table></div>
</details>
<details class='section-fold'><summary class='section-sum section-sum-row'><span class='sum-main'>(3) Unmatched — {len(unmatched)} legs เรียงตามเวลา</span><span class='sum-dl'>{_xlsx_dl('03_Unmatched_Legs.xlsx', 'ตาราง (3)')}</span></summary>
<div class='panel'>
<p class='sub'>UM-O = Origin ไม่มี Dest คู่ · UM-D = Dest ไม่มี Origin คู่ · max_travel_h={cfg.max_travel_h}h · match เลือก Origin ที่ t_in ล่าสุดก่อน Dest · <b>อยู่จุด (ชม.)</b> = เวลาระหว่างเข้า–ออกของแถวนั้น · <b>ถึงเข้าครั้งถัดไป</b> = จากเวลาออกของแถวนี้ถึงเวลาเข้าของเหตุการณ์ถัดไป (เรียงทะเบียนเดียวกัน จากไฟล์ Origin+Dest)</p>
<table><thead><tr><th>ประเภท</th><th>ทะเบียน</th><th>เวลาเข้า</th><th>เวลาออก</th><th>อยู่จุด (ชม.)</th><th>ถึงเข้าครั้งถัดไป (ชม.)</th><th>เหตุผล</th></tr></thead><tbody>
{um_section_html}
</tbody></table></div>
</details>
<details class='section-fold'><summary class='section-sum section-sum-row'><span class='sum-main'>(คลิกเพื่อขยาย) Audit Log — เหตุผลการคิดเงิน รายวัน × ทะเบียน</span><span class='sum-dl'>{_xlsx_dl('04_Audit_Log.xlsx', 'Audit')}</span></summary>
<div class='panel'><p class='sub'>ทุกแถวอธิบายว่าวันนั้นทะเบียนนั้นคิดเงินอย่างไร</p>
<table><thead><tr><th>{'วันงาน' if cfg.use_origin_day_fifty else 'วันที่ Dest_In'}</th><th>ทะเบียน</th><th>Site</th><th>เที่ยว</th><th>เรท(฿)</th><th>ค่าเที่ยว(฿)</th><th>ส่วนเพิ่ม (฿)</th><th>ขากลับ(฿)</th><th>รวม(฿)</th><th>เหตุผล</th></tr></thead><tbody>
{audit_html or "<tr><td colspan=10>ไม่มีข้อมูล</td></tr>"}
</tbody></table></div></details>
<details class='section-fold'><summary class='section-sum section-sum-row'><span class='sum-main'>รายทะเบียน</span><span class='sum-dl'>{_xlsx_dl('02_Plate_DestDay_Daily.xlsx', 'เดลี่×ทะเบียน')}</span></summary>
<div class='panel'><ul>{''.join(f"<li><a href='plates/{esc(p)}.html'>{esc(p)}</a></li>" for p in plates)}</ul></div>
</details></body></html>"""

    report_dir.mkdir(parents=True, exist_ok=True)
    (report_dir / "index.html").write_text(idx, encoding="utf-8")

    _trips_plate_opts = "".join(f"<option value='{esc(p)}'>{esc(p)}</option>" for p in plates)
    trips_html_content = (
        f"""<!doctype html><html><head><meta charset='utf-8'><meta name='viewport' content='width=device-width,initial-scale=1'>
<title>Trips</title><style>{css}</style></head><body>
<div class='h1'>เที่ยวทั้งหมด <span class='trips-tag'>หน้าหลักลูกค้า</span></div>
<div class='trips-lead'>เวลาเข้า-ออกครบทุกขา · ค่าขนส่ง / เสียเวลา / ขากลับ — กรองทะเบียนได้ด้านล่าง</div>
<div class='nav'><a href='index.html'>&larr; สรุปภาพรวม</a> · <a href='../../../Oatside/Oatside_PG_Trip_Summary_By_Site.xlsx'>Excel รวมทุกชีต</a></div>
<div class='panel'><div class='panel-title-row'><h3>เที่ยวทั้งหมด (matched + unmatched)</h3><a class='xlsx-dl' href='exports/05_Trip_Detail.xlsx' download onclick='event.stopPropagation()'>ดาวน์โหลด Excel (Trip Detail)</a></div>
<p class='sub'>เรียงตามเวลา (matched ใช้ Origin In · unmatched ใช้เวลาขา Origin/Destination) — UM-O/UM-D เว้นฝั่งที่ยังไม่มีคู่เป็น —<br>
<b>ค่าเงิน:</b> ค่าขนส่ง = เรทวัน Dest_In ของเที่ยวนั้น · <b>เสียเวลา+50%/+100%</b> = ยอดรวมส่วนเพิ่ม fifty ของ (ทะเบียน×วัน Dest_In) แสดงที่แถวแรกของวันนั้น — <b>ไม่ได้คิดจากชั่วโมงในช่อง Dest Wait โดยตรง</b> (สีส้ม = แค่เตือนว่ารอปลายทางเกินเกณฑ์) · <b>ขากลับ(฿)</b> = ยอดจาก <code>manual_return_trips</code> แสดงที่แถวแรกของวันนั้น (ไม่เพิ่มจำนวนเที่ยว matched)</p>
<div class='filter-bar'><label for='tripsPlateFilter'>กรองทะเบียน</label><select id='tripsPlateFilter'><option value=''>ทุกคัน</option>{_trips_plate_opts}</select><label for='tripsPlateQuery' style='margin-left:6px'>ค้นหา</label><input id='tripsPlateQuery' type='search' placeholder='พิมพ์ค้นหา...' autocomplete='off'></div>
<div class='table-scroll'><table id='tripsAllTable'><thead><tr><th>Origin Date</th><th>Dest Date</th><th>Site</th><th>ทะเบียน</th><th>Origin In</th><th>Origin Out</th><th>Dest In</th><th>Dest Out</th><th>Orig Wait</th><th>Travel</th><th>Dest Wait</th><th>อยู่จุด UM (ชม.)</th><th>ถึงเข้าครั้งถัดไป (ชม.)</th><th>ค่าขนส่ง(฿)</th><th>เสียเวลา+50%(฿)</th><th>เสียเวลา+100%(฿)</th><th>ตีเปล่า+50%(฿)</th><th>ขากลับ(฿)</th></tr></thead><tbody>
{merged_all_rows}
</tbody></table></div></div>
"""
        + _TRIPS_FILTER_JS
        + "\n</body></html>"
    )

    (report_dir / "trips.html").write_text(trips_html_content, encoding="utf-8")

    plates_dir = report_dir / "plates"
    plates_dir.mkdir(exist_ok=True)
    for old in plates_dir.glob("*.html"):
        old.unlink()
    by_plate: dict[str, list[Trip]] = defaultdict(list)
    for t in trips:
        by_plate[t.plate].append(t)
    # build audit note index by (plate, origin_day) for plate-page reason display
    audit_oday_idx: dict[tuple[str, date], str] = {}
    if cfg.use_origin_day_fifty:
        for _ar in audit_rows:
            if "origin_day" in _ar:
                audit_oday_idx[(_ar["plate"], _ar["origin_day"])] = _ar["billing_note"]
    for p, lst in by_plate.items():
        if cfg.use_origin_day_fifty:
            by_oday: dict[date, list[Trip]] = defaultdict(list)
            for t in lst:
                by_oday[t.o_in.date()].append(t)
            day_rows = []
            for od in sorted(by_oday.keys()):
                cnt = len(by_oday[od])
                frs = fifty_origin_lists.get((p, od), [])
                reason = audit_oday_idx.get((p, od), f"ไม่เก็บ (วันงาน {cnt} เที่ยว)" if cnt != 1 else "ไม่เก็บ (override หรือเงื่อนไขเพิ่ม)")
                badge = ""
                if frs:
                    parts = [html_fifty_surcharge_badge(x, cfg) for x in frs if int(x.get("surcharge_baht", 0) or 0) > 0]
                    parts = [b for b in parts if b]
                    if parts:
                        badge = " " + " ".join(parts)
                elif cnt == 1:
                    badge = " <span class='badge lcb'>1 เที่ยว (ไม่เก็บ)</span>"
                nw_sum = sum(trip_no_work_outbound_baht(t, first_no_work, cfg) for t in by_oday[od])
                nw_cell = f"฿{nw_sum:,}" if nw_sum else "—"
                day_rows.append(
                    f"<tr><td>{od}</td><td>{cnt}</td><td>{badge}</td>"
                    f"<td class='note'>{esc(reason)}</td><td>{nw_cell}</td></tr>"
                )
            day_tbl = "".join(day_rows)
            summary_hdr = "รายวันงาน (Origin_In)"
            summary_sub = "<p class='sub'>วันงาน = วันที่ Origin_In · ข้ามคืนไม่แตกวัน · วันไม่มี Origin ไม่นับ</p>"
            day_thead = "<tr><th>วันงาน</th><th>เที่ยว</th><th>ส่วนเพิ่ม</th><th>เหตุผล</th><th>ตีเปล่า+50%(฿)</th></tr>"
        else:
            by_day: dict[date, list[Trip]] = defaultdict(list)
            for t in lst:
                by_day[t.dest_date].append(t)
            day_rows = []
            for d in sorted(by_day.keys()):
                cnt = len(by_day[d])
                frs = fifty_by_lists.get((p, d), [])
                badge = ""
                if frs:
                    parts = [html_fifty_surcharge_badge(x, cfg) for x in frs if int(x.get("surcharge_baht", 0) or 0) > 0]
                    parts = [b for b in parts if b]
                    if parts:
                        badge = " " + " ".join(parts)
                elif cnt == 1:
                    badge = " <span class='badge lcb'>1 เที่ยว (ไม่เก็บ +%)</span>"
                day_rows.append(f"<tr><td>{d}</td><td>{cnt}</td><td>{badge}</td></tr>")
            day_tbl = "".join(day_rows)
            summary_hdr = "รายวัน (Dest_In)"
            summary_sub = ""
            day_thead = "<tr><th>วันที่</th><th>เที่ยว</th><th>หมายเหตุ billing</th></tr>"
        merged_plate_rows = interleaved_matched_unmatched_rows_html(
            lst,
            unmatched,
            trip_row_plate,
            plate=p,
            include_plate_link=False,
            include_plate_column=False,
            leg_timeline_by_plate=leg_timeline_by_plate,
        )
        pg = f"""<!doctype html><html><head><meta charset='utf-8'><meta name='viewport' content='width=device-width,initial-scale=1'>
<title>{esc(p)}</title><style>{css}</style></head><body>
<div class='h1'>ทะเบียน {esc(p)}</div>
<div class='nav'><a href='../index.html'>&larr; กลับสรุป</a> | <a href='../trips.html'>ดูเที่ยวทั้งหมด</a></div>
<div class='panel'><h3>{summary_hdr}</h3>{summary_sub}<table><thead>{day_thead}</thead><tbody>{day_tbl}</tbody></table></div>
<div class='panel'><h3>รายเที่ยว (matched + unmatched)</h3>
<p class='sub'>เรียงตามเวลา (matched ใช้ Origin In · unmatched ใช้เวลาขา Origin/Destination) — UM-O/UM-D เว้นฝั่งที่ยังไม่มีคู่เป็น —<br>หัวตารางล่างเลื่อนตามแบบ freeze แถว (เลื่อนในกรอบ)</p>
<div class='table-scroll'><table><thead><tr><th>Origin Date</th><th>Dest Date</th><th>Site</th><th>Origin In</th><th>Origin Out</th><th>Dest In</th><th>Dest Out</th><th>Orig Wait</th><th>Travel</th><th>Dest Wait</th><th>อยู่จุด UM (ชม.)</th><th>ถึงเข้าครั้งถัดไป (ชม.)</th><th>ค่าขนส่ง(฿)</th><th>เสียเวลา+50%(฿)</th><th>เสียเวลา+100%(฿)</th><th>ตีเปล่า+50%(฿)</th><th>ขากลับ(฿)</th></tr></thead><tbody>{merged_plate_rows}</tbody></table></div></div>
</body></html>"""
        (plates_dir / f"{p}.html").write_text(pg, encoding="utf-8")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    cfg = load_oatside_config()
    folder = _oatside_dir()
    origin_path, dest_path = discover_gps_files(folder)
    trips, unmatched, _travels = build_trips(origin_path, dest_path, cfg)
    daily_rows = daily_activity_by_dest(trips, cfg)
    daily_time = daily_time_rows(trips, unmatched, cfg)
    actual, commit, short, extra = billing_totals(daily_rows, cfg)
    bc_stats = site_billing(daily_rows, cfg)
    overrides = load_billing_overrides()
    if cfg.use_origin_day_fifty:
        fifty_rows, fifty_total = one_trip_fifty_pct_origin_day(trips, overrides, cfg)
    elif cfg.use_origin_24h_fifty:
        fifty_rows, fifty_total = one_trip_fifty_pct_details_origin24h(trips, overrides, cfg)
    else:
        fifty_rows, fifty_total = one_trip_fifty_pct_details(trips, overrides, cfg)
    add_fr, add_tot = supplement_long_dest_wait_midnight_fifty(trips, fifty_rows, overrides, cfg)
    fifty_rows = fifty_rows + add_fr
    fifty_total += add_tot
    if cfg.use_origin_day_fifty:
        audit_rows = origin_day_audit_rows(trips, fifty_rows, overrides, cfg)
    elif cfg.use_origin_24h_fifty:
        audit_rows = audit_log_rows(trips, fifty_rows, overrides, cfg)
    else:
        audit_rows = audit_log_rows(trips, fifty_rows, overrides, cfg)
    merge_manual_extra_into_audit(audit_rows, cfg)
    merge_manual_return_into_audit(audit_rows, cfg)
    base_baht = base_trips_revenue_baht(trips, cfg) + sum_manual_extra_baht(cfg)
    o_legs_all = parse_legs(origin_path)
    leg_timeline_by_plate = build_leg_timeline_by_plate(o_legs_all, parse_legs(dest_path))
    nw_rows, nw_total = no_work_outbound_rows(trips, cfg)
    pday_rows = plate_dest_day_rows(trips, fifty_rows, cfg, nw_rows=nw_rows)
    merge_manual_extra_into_pday(pday_rows, cfg)
    merge_manual_return_into_pday(pday_rows, cfg)
    min_trip_money = int(extra) if cfg.charge_min_trip_shortfall else 0
    if not cfg.charge_min_trip_shortfall:
        bc_a, bc_c, bc_s, bc_e, lc_a, lc_c, lc_s, lc_e = bc_stats
        bc_stats = (bc_a, bc_c, bc_s, 0, lc_a, lc_c, lc_s, 0)
    phantom_rows = phantom_zero_trip_candidates(o_legs_all, trips, cfg)
    hint_rows = double_origin_um_hints(unmatched)
    grand_extra = min_trip_money + int(fifty_total) + int(nw_total)
    customer_grand_baht = int(base_baht) + int(grand_extra) + int(sum_manual_return_baht(cfg))

    cpd_rows = customer_trips_per_day_rows(trips)
    apply_manual_extra_to_cpd(cpd_rows, cfg)

    xlsx_out = folder / "Oatside_PG_Trip_Summary_By_Site.xlsx"
    write_excel(
        xlsx_out,
        origin_path.name,
        dest_path.name,
        trips,
        unmatched,
        daily_time,
        daily_rows,
        fifty_rows,
        int(fifty_total),
        min_trip_money,
        audit_rows,
        cfg,
        int(customer_grand_baht),
        nw_rows,
        int(nw_total),
        phantom_rows,
        hint_rows,
        pday_rows,
        cpd_rows,
        leg_timeline_by_plate,
    )
    report_dir = _root() / "TransportRateCalculator" / "reports" / "oatside-apr2026"
    write_split_excel_exports(
        xlsx_out,
        report_dir,
        built_at=datetime.now().strftime("%Y-%m-%d %H:%M"),
    )

    write_html(
        report_dir,
        origin_path.name,
        trips,
        daily_rows,
        daily_time,
        actual,
        commit,
        short,
        min_trip_money,
        bc_stats,
        fifty_rows,
        int(fifty_total),
        grand_extra,
        int(base_baht),
        int(customer_grand_baht),
        pday_rows,
        audit_rows,
        unmatched,
        int(nw_total),
        cfg,
        cpd_rows,
        leg_timeline_by_plate,
    )

    print(f"Config:  {_config_path()}")
    print(f"Trips: {len(trips)} | Unmatched legs: {len(unmatched)}")
    print(f"Excel:   {xlsx_out}")
    print(f"HTML:    {report_dir}")


if __name__ == "__main__":
    main()
