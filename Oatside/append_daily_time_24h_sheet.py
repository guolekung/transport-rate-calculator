"""
Post-process Oatside_PG_Trip_Summary_By_Site.xlsx:
add sheet Daily_Time_24h_Check (matched cycle + unmatched dwell vs 24h gap).

Prefer: python Oatside/build_oatside_reports.py (rebuilds from GPS + writes this sheet).

This script is for older workbooks that already have Trip_Detail + Unmatched_Log.
Run from repo root:  python Oatside/append_daily_time_24h_sheet.py
"""

from __future__ import annotations

import re
import sys
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

import openpyxl

BIGC_EXACT = frozenset({"71-5041", "71-5042"})


def site_for_plate(plate: str) -> str:
    p = (plate or "").strip()
    if p in BIGC_EXACT:
        return "BigC"
    m = re.match(r"^71-(\d+)$", p)
    if m:
        n = int(m.group(1))
        if 8000 <= n <= 8009:
            return "BigC"
    return "LCB"


def _to_date(val) -> date | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        s = val.strip()
        for fmt, ln in (("%Y-%m-%d %H:%M:%S", 19), ("%Y-%m-%d", 10)):
            try:
                return datetime.strptime(s[:ln], fmt).date()
            except ValueError:
                continue
    return None


def _to_dt(val) -> datetime | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, date):
        return datetime.combine(val, datetime.min.time())
    if isinstance(val, str):
        s = val.strip()[:19]
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                continue
        try:
            return datetime.fromisoformat(s.replace("Z", "+00:00"))
        except ValueError:
            return None
    return None


def hours_between(a: datetime, b: datetime) -> float:
    return (b - a).total_seconds() / 3600.0


def workbook_path() -> Path:
    here = Path(__file__).resolve().parent
    p = here / "Oatside_PG_Trip_Summary_By_Site.xlsx"
    if p.exists():
        return p
    # run from repo root
    p2 = here.parent / "Oatside" / "Oatside_PG_Trip_Summary_By_Site.xlsx"
    if p2.exists():
        return p2
    raise FileNotFoundError("Oatside_PG_Trip_Summary_By_Site.xlsx not found")


def main() -> None:
    path = workbook_path()
    wb = openpyxl.load_workbook(path)

    trip_ws = wb["Trip_Detail"]
    headers: dict[str, int] = {}
    for c in trip_ws[1]:
        if c.value:
            headers[str(c.value)] = c.column

    col_plate = headers["Plate"]
    col_trip_date = headers["Trip_Date"]
    col_total = headers["Total_Cycle_h"]

    matched: dict[tuple[date, str], float] = defaultdict(float)
    for row in range(2, trip_ws.max_row + 1):
        plate = trip_ws.cell(row=row, column=col_plate).value
        td = trip_ws.cell(row=row, column=col_trip_date).value
        tot = trip_ws.cell(row=row, column=col_total).value
        if not plate or tot is None:
            continue
        d = _to_date(td)
        if not d:
            continue
        try:
            h = float(tot)
        except (TypeError, ValueError):
            continue
        matched[(d, str(plate).strip())] += h

    um_ws = wb["Unmatched_Log"]
    um_h: dict[str, int] = {}
    for c in um_ws[1]:
        if c.value:
            um_h[str(c.value)] = c.column
    c_src = um_h.get("Source")
    c_plate = um_h["Plate"]
    c_in = um_h["In"]
    c_out = um_h["Out"]

    um_origin: dict[tuple[date, str], float] = defaultdict(float)
    um_dest: dict[tuple[date, str], float] = defaultdict(float)
    for row in range(2, um_ws.max_row + 1):
        plate = um_ws.cell(row=row, column=c_plate).value
        t_in = _to_dt(um_ws.cell(row=row, column=c_in).value)
        t_out = _to_dt(um_ws.cell(row=row, column=c_out).value)
        if not plate or not t_in or not t_out:
            continue
        src = um_ws.cell(row=row, column=c_src).value if c_src else None
        src_s = str(src or "")
        d = t_in.date()
        h = hours_between(t_in, t_out)
        if h < 0 or h > 72:
            continue
        key = (d, str(plate).strip())
        if "Origin" in src_s:
            um_origin[key] += h
        elif "Destination" in src_s or "Dest" in src_s:
            um_dest[key] += h
        else:
            um_origin[key] += h

    keys = sorted(set(matched) | set(um_origin) | set(um_dest), key=lambda x: (x[0], x[1]))

    sheet_name = "Daily_Time_24h_Check"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.append(
        [
            "Activity_Date",
            "Plate",
            "Site",
            "Matched_Cycle_h",
            "Unmatched_Origin_h",
            "Unmatched_Dest_h",
            "Unmatched_Total_h",
            "Combined_h",
            "Gap_vs_24_h",
        ]
    )
    for d, plate in keys:
        mk = (d, plate)
        m = matched.get(mk, 0.0)
        uo = um_origin.get(mk, 0.0)
        ud = um_dest.get(mk, 0.0)
        ut = uo + ud
        comb = m + ut
        gap = 24.0 - comb
        ws.append(
            [
                d.isoformat(),
                plate,
                site_for_plate(plate),
                round(m, 2),
                round(uo, 2),
                round(ud, 2),
                round(ut, 2),
                round(comb, 2),
                round(gap, 2),
            ]
        )

    info = wb["Info"]
    r = info.max_row + 2
    info.cell(row=r, column=1, value="Daily_Time_24h_Check")
    info.cell(
        row=r + 1,
        column=1,
        value=(
            "Matched_Cycle_h: sum of Total_Cycle_h from Trip_Detail grouped by Trip_Date + Plate. "
            "Unmatched_*: sum of (Out-In) from Unmatched_Log by calendar date of In + Plate; split Origin/Dest. "
            "Combined_h = Matched + Unmatched_Total. Gap_vs_24_h = 24 - Combined (remaining time not from these logs)."
        ),
    )

    try:
        wb.save(path)
        out = path
    except PermissionError:
        alt = path.with_name(path.stem + "_Daily24h.xlsx")
        wb.save(alt)
        out = alt
        print(f"NOTE: {path.name} is open/locked; saved to {alt.name} instead.", file=sys.stderr)
    print(f"Updated {out} sheet {sheet_name} ({len(keys)} rows).")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(e, file=sys.stderr)
        sys.exit(1)
